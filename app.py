import os
import json
import base64
import threading
import datetime as dt
import time
import signal
import sys
import subprocess
import io
import csv
from typing import Dict, Any, Optional, List
import numpy as np
import cv2
import shutil
import socket
import ctypes
import ctypes.wintypes as wintypes
from cachetools import TTLCache
from openpyxl import Workbook
from openpyxl.styles import Font

from flask import Flask, jsonify, render_template, request, send_file, Response, send_from_directory
from flask_socketio import SocketIO, emit
from database_models import SessionLocal, Employee, Camera, FaceTemplate, Attendance, Presence, Event, AlertLog, init_db
from database_models import seed_cameras_from_configs

# --- System Uptime Tracking ---
_system_start_time = None

# --- TensorRT Engine Cache Configuration ---
# Set a dedicated cache directory to prevent cluttering the root folder.
# This must be done BEFORE module_AI is imported.
TENSORRT_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_tensorrt_cache')
os.makedirs(TENSORRT_CACHE_DIR, exist_ok=True)
os.environ['ORT_TENSORRT_CACHE_PATH'] = TENSORRT_CACHE_DIR

# AI tracking manager
try:
    from module_AI import ai_manager
except Exception as _e:
    ai_manager = None

import cv2 # cv2 sudah di-import di banyak tempat, pastikan terinstall

# Inisialisasi Flask dan Socket.IO
app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = 'change-me'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Inisialisasi variabel global
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CAMERA_DIR = os.path.join(BASE_DIR, 'camera_configs')
DB_DIR = os.path.join(BASE_DIR, 'db')
os.makedirs(DB_DIR, exist_ok=True)
FACE_IMG_DIR = os.path.join(BASE_DIR, 'face_images')
DB_PATH = os.path.join(BASE_DIR, 'db', 'attendance.db')
CAPTURE_DIR = os.path.join(BASE_DIR, 'captures')
os.makedirs(CAPTURE_DIR, exist_ok=True)
TRACK_STATE_PATH = os.path.join(BASE_DIR, 'config', 'tracking_mode.json')
PARAMS_PATH = os.path.join(BASE_DIR, 'config', 'parameter_config.json')
_camera_map: Dict[int, Dict[str, Any]] = {}
_app_params: Dict[str, Any] = {}

# --- Captures background saver (per-camera rolling cache) ---
CAPTURES_DIR = os.path.join(BASE_DIR, 'captures')
os.makedirs(CAPTURES_DIR, exist_ok=True)
_saver_thread_started = False

# Dedicated Attendance Captures dir
ATT_CAPTURES_DIR = os.path.join(BASE_DIR, 'attendance_captures')
os.makedirs(ATT_CAPTURES_DIR, exist_ok=True)

def _list_camera_status() -> List[Dict[str, Any]]:
    """Return camera status list similar to /api/cameras/status items.
    Falls back to configs with stream_enabled when status API unavailable."""
    try:
        import urllib.request
        req = urllib.request.Request('http://127.0.0.1:5000/api/cameras/status')
        with urllib.request.urlopen(req, timeout=2.5) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            items = data.get('items') if isinstance(data, dict) else None
            if isinstance(items, list):
                return items
    except Exception:
        pass
    # Fallback: load from configs and assume stream_enabled as activity
    cams = load_cameras()
    out = []
    for cid, meta in cams.items():
        out.append({'id': cid, 'ai_running': False, 'stream_enabled': bool(meta.get('stream_enabled', False))})
    return out

def _save_snapshot_for_camera(cam_id: int) -> bool:
    """Fetch snapshot from our own API and store to captures/<cam_id>/; return True if saved."""
    try:
        import urllib.request
        # Use WIB timezone for filename timestamp (consistent with system)
        ts = _now_wib().strftime('%Y%m%d_%H%M%S')
        url = f'http://127.0.0.1:5000/api/cameras/{int(cam_id)}/snapshot?annotate=1'
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=4.0) as resp:
            data = resp.read()
        cam_dir = os.path.join(CAPTURES_DIR, str(int(cam_id)))
        os.makedirs(cam_dir, exist_ok=True)
        fname = f'{ts}.jpg'
        fpath = os.path.join(cam_dir, fname)
        with open(fpath, 'wb') as f:
            f.write(data)
        # Rotate to max 5 files
        try:
            files = sorted([x for x in os.listdir(cam_dir) if x.lower().endswith('.jpg')])
            if len(files) > 5:
                for old in files[0:len(files)-5]:
                    try: os.remove(os.path.join(cam_dir, old))
                    except Exception: pass
        except Exception:
            pass
        print(f"[Frame Capture] Saved snapshot for CAM{cam_id}: {fname}")
        return True
    except Exception as e:
        print(f"[Frame Capture] Failed to save snapshot for CAM{cam_id}: {e}")
        return False

def _background_capture_saver_loop(interval_sec: int = 5):
    while True:
        try:
            items = _list_camera_status()
            for it in items:
                try:
                    if not it: continue
                    if not (it.get('ai_running') or it.get('stream_enabled')):
                        continue
                    _save_snapshot_for_camera(int(it.get('id')))
                except Exception:
                    continue
        except Exception:
            pass
        time.sleep(max(1, int(interval_sec)))

def _ensure_saver_started():
    global _saver_thread_started
    if _saver_thread_started:
        return
    t = threading.Thread(target=_background_capture_saver_loop, args=(5,), daemon=True)
    t.start()
    _saver_thread_started = True

def _start_capture_saver_thread():
    _ensure_saver_started()

_workers_by_sid: Dict[str, Any] = {}
# Camera status cache with bounded size and TTL to prevent memory leaks
_CAM_STATUS_TTL = 10.0  # seconds
_cam_status_cache: TTLCache = TTLCache(maxsize=100, ttl=_CAM_STATUS_TTL)
# Optional: Face embedding engine (lazy)
_face_app = None
# Directory to store cropped face images per template
FACE_IMG_DIR = os.path.join(BASE_DIR, 'face_images')


def _safe_name(name: Optional[str]) -> str:
    try:
        s = (name or '').strip()
        if not s:
            return 'unknown'
        # Keep alnum, space, dash, underscore; replace space with underscore
        cleaned = ''.join(ch for ch in s if ch.isalnum() or ch in (' ', '-', '_'))
        cleaned = '_'.join(part for part in cleaned.split())
        return cleaned[:64]  # limit length
    except Exception:
        return 'unknown'

def _employee_image_dir(emp: 'Employee') -> str:
    return os.path.join(FACE_IMG_DIR, _safe_name(getattr(emp, 'name', None)) or f"ID_{emp.id}")

# --- Time helpers ---
def _to_iso_utc(dtobj: Optional[dt.datetime]) -> Optional[str]:
    """Serialize a datetime as ISO 8601.
    If timezone-aware (WIB/UTC+7), return with timezone offset.
    If naive, assume WIB and return with +07:00 offset.
    """
    if dtobj is None:
        return None
    try:
        if dtobj.tzinfo is None:
            # Assume naive timestamps are WIB (UTC+7)
            wib_tz = dt.timezone(dt.timedelta(hours=7))
            dtobj = dtobj.replace(tzinfo=wib_tz)
        return dtobj.isoformat()
    except Exception:
        try:
            return dtobj.isoformat()
        except Exception:
            return None


def _to_wib_string(dtobj: Optional[dt.datetime]) -> Optional[str]:
    """Convert UTC datetime to WIB (UTC+7) and format as readable string.
    Database stores timestamps as UTC-naive; treat as UTC and convert to WIB.
    Returns: 'YYYY-MM-DD HH:MM:SS' in WIB timezone
    """
    if dtobj is None:
        return None
    try:
        # Treat naive datetime as UTC
        if dtobj.tzinfo is None:
            dtobj_utc = dtobj.replace(tzinfo=dt.timezone.utc)
        else:
            dtobj_utc = dtobj.astimezone(dt.timezone.utc)

        # Convert to WIB (UTC+7)
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        dtobj_wib = dtobj_utc.astimezone(wib_tz)

        # Format as string without timezone info
        return dtobj_wib.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        # Fallback to ISO format if conversion fails
        try:
            return dtobj.isoformat(sep=' ')
        except Exception:
            return None


def _get_face_app():
    global _face_app
    if _face_app is not None:
        return _face_app
    try:
        from insightface.app import FaceAnalysis
        # Use the same providers and detection size as module_AI (from parameter_config.json)
        det_size = (640, 640)
        providers = ['CPUExecutionProvider']
        try:
            if ai_manager is not None and hasattr(ai_manager, 'engine') and ai_manager.engine is not None:
                det_size = tuple(ai_manager.engine.det_size)
                providers = list(ai_manager.engine.providers)
        except Exception:
            pass
        # Try with configured providers first, then fallback to defaults
        try:
            # Tentukan ctx_id=0 (GPU) hanya jika provider GPU/TensorRT ada.
            ctx = 0 if any('CUDA' in p or 'Tensorrt' in p for p in providers) else -1
            
            _face_app = FaceAnalysis(name='buffalo_l', providers=providers)
            _face_app.prepare(ctx_id=ctx, det_size=det_size) # Menggunakan ctx yang sudah ditentukan
            return _face_app
        except Exception as e1:
            print(f"FaceAnalysis init with providers {providers} failed: {e1}. Retrying with default providers...")
            # Saat fallback, pastikan tidak ada panggilan ke GPU
            _face_app = FaceAnalysis(name='buffalo_l')
            _face_app.prepare(ctx_id=-1, det_size=det_size) # Gunakan ctx_id=-1 (CPU)
            return _face_app
    except Exception as e:
        print(f"Failed to init FaceAnalysis: {e}")
        return None


def load_cameras() -> Dict[int, Dict[str, Any]]:
    cams: Dict[int, Dict[str, Any]] = {}
    if not os.path.isdir(CAMERA_DIR):
        return cams
    for name in os.listdir(CAMERA_DIR):
        path = os.path.join(CAMERA_DIR, name)
        if not os.path.isdir(path):
            continue
        cfg_path = os.path.join(path, 'config.json')
        if not os.path.isfile(cfg_path):
            continue
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            cam_id = int(cfg.get('id'))
            cams[cam_id] = {
                'id': cam_id,
                'name': cfg.get('name', f'CAM {cam_id}'),
                'rtsp_url': cfg.get('rtsp_url', ''),
                'enabled': bool(cfg.get('enabled', False)),  # AI enabled flag (legacy)
                'stream_enabled': bool(cfg.get('stream_enabled', True)),
                'area': cfg.get('area') or cfg.get('zone') or '',
            }
        except Exception:
            continue
    return cams

# --- App parameters ---
def load_params() -> Dict[str, Any]:
    global _app_params
    if _app_params: # Return cached if already loaded
        return _app_params
    try:
        if os.path.isfile(PARAMS_PATH):
            with open(PARAMS_PATH, 'r', encoding='utf-8') as f:
                _app_params = json.load(f) or {}
    except Exception:
        _app_params = {}
    # Defaults
    _app_params.setdefault('away_mute_threshold_hours', 15)
    _app_params.setdefault('mark_absent_enabled', True)
    _app_params.setdefault('mark_absent_offset_minutes_before_end', 5)
    return _app_params


# --- Daily maintenance: purge events older than today ---

def purge_old_events():
    try:
        with SessionLocal() as db:
            from database_models import Event  # local import to avoid circular
            # Compute start of today in server local time converted to UTC assumption-free by comparing date only
            today = dt.date.today()
            # Delete events where date(timestamp) != today
            # SQLite lacks date() on SQLAlchemy by default; fetch and delete in chunks
            rows = db.query(Event).all()
            removed = 0
            for r in rows:
                try:
                    if not r.timestamp or r.timestamp.date() != today:
                        db.delete(r)
                        removed += 1
                except Exception:
                    continue
            db.commit()
            if removed:
                print(f"[MAINT] Purged {removed} old events (kept only today)")
    except Exception as e:
        print(f"[MAINT] purge_old_events error: {e}")


def _seconds_until_midnight_local() -> float:
    now = _now_local()
    tomorrow = now.date() + dt.timedelta(days=1)
    midnight = dt.datetime.combine(tomorrow, dt.time.min, tzinfo=now.tzinfo)
    return max(1.0, (midnight - now).total_seconds())


def schedule_midnight_purge():
    def _job():
        while True:
            try:
                purge_old_events()
            except Exception:
                pass
            time.sleep(_seconds_until_midnight_local())
    t = threading.Thread(target=_job, daemon=True)
    t.start()

# Initialize schedulers at import time (single-process assumption)
# Moved to after function definitions below to avoid NameError on import ordering.

# --- Daily retention: cleanup old attendance captures ---
def _cleanup_old_attendance_captures():
    try:
        keep_days = int(_app_params.get('attendance_captures_retention_days') or 30)
        keep_days = max(1, min(3650, keep_days))
        cutoff = dt.date.today() - dt.timedelta(days=keep_days)
        root = ATT_CAPTURES_DIR
        if not os.path.isdir(root):
            return
        removed = 0
        for name in os.listdir(root):
            path = os.path.join(root, name)
            if not os.path.isdir(path):
                continue
            try:
                folder_date = dt.datetime.strptime(name, '%Y-%m-%d').date()
            except Exception:
                continue
            if folder_date < cutoff:
                try:
                    shutil.rmtree(path, ignore_errors=True)
                    removed += 1
                except Exception:
                    pass
        if removed:
            print(f"[RETENTION] Removed {removed} attendance_captures folder(s) older than {keep_days} days")
    except Exception as e:
        print(f"[RETENTION] cleanup error: {e}")

def schedule_attendance_retention():
    def _job():
        while True:
            try:
                _cleanup_old_attendance_captures()
            except Exception:
                pass
            time.sleep(_seconds_until_midnight_local())
    t = threading.Thread(target=_job, daemon=True)
    t.start()

# --- Absent Employee Detection ---
def _mark_absent_employees():
    """Mark all active employees without attendance today as ABSENT.
    Runs daily at 17:30 WIB (end of work hours).
    Skips manual entries (entry_type='MANUAL') to preserve admin overrides.
    """
    try:
        with SessionLocal() as db:
            today = dt.date.today()
            # Get all active employees
            active_employees = db.query(Employee).filter(Employee.is_active == True).all()
            marked_count = 0
            skipped_manual = 0

            for emp in active_employees:
                # Check if employee has attendance record for today
                attendance = db.query(Attendance).filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date == today
                ).first()

                if attendance is None:
                    # No attendance record, create ABSENT entry with SYSTEM type
                    new_attendance = Attendance(
                        employee_id=emp.id,
                        date=today,
                        status='ABSENT',
                        first_in_ts=None,
                        last_out_ts=None,
                        entry_type='SYSTEM'
                    )
                    db.add(new_attendance)
                    marked_count += 1
                elif attendance.status != 'ABSENT' and attendance.first_in_ts is None:
                    # Has record but no first_in_ts
                    # Skip if manual entry (admin has manually set status)
                    if attendance.entry_type == 'MANUAL':
                        skipped_manual += 1
                        continue
                    # Otherwise update to ABSENT
                    attendance.status = 'ABSENT'
                    attendance.entry_type = 'SYSTEM'
                    marked_count += 1

            db.commit()
            if marked_count > 0:
                print(f"[ABSENT DETECTION] Marked {marked_count} employee(s) as ABSENT for {today}")
            if skipped_manual > 0:
                print(f"[ABSENT DETECTION] Skipped {skipped_manual} manual entry(ies)")
    except Exception as e:
        print(f"[ABSENT DETECTION] Error: {e}")

def _seconds_until_target_time(target_hour: int, target_minute: int) -> float:
    """Calculate seconds until next occurrence of target time (WIB)."""
    now = _now_local()
    target = now.replace(hour=target_hour, minute=target_minute, second=0, microsecond=0)

    # If target time already passed today, schedule for tomorrow
    if now >= target:
        target = target + dt.timedelta(days=1)

    seconds = (target - now).total_seconds()
    return max(1.0, seconds)

def schedule_absent_detection():
    """Schedule daily absent employee detection at 17:30 WIB."""
    def _job():
        while True:
            try:
                # Wait until 17:30 WIB
                wait_seconds = _seconds_until_target_time(17, 30)
                print(f"[ABSENT DETECTION] Scheduled for {wait_seconds/3600:.1f} hours from now")
                time.sleep(wait_seconds)

                # Run absent detection
                _mark_absent_employees()
            except Exception as e:
                print(f"[ABSENT DETECTION] Scheduler error: {e}")
                # On error, wait 1 hour before retry
                time.sleep(3600)

    t = threading.Thread(target=_job, daemon=True)
    t.start()
    print("[ABSENT DETECTION] Scheduler started (daily at 17:30 WIB)")

# --- Tracking schedule & state (WIB) ---
def _default_tracking_state():
    return {
        'auto_schedule': True,
        'tracking_active': False,   # will be computed at startup
        'suppress_alerts': False,
        'pause_until': None,        # ISO local time string or None
        'work_hours': '08:30-17:30',
        'lunch_break': '12:00-13:00',
    }

_tracking_state = _default_tracking_state()

def _load_tracking_state():
    global _tracking_state
    try:
        if os.path.isfile(TRACK_STATE_PATH):
            with open(TRACK_STATE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    _tracking_state.update(data)
                    # Validate and clear expired pauses on load
                    pause_until = _tracking_state.get('pause_until')
                    if pause_until:
                        try:
                            until = dt.datetime.fromisoformat(pause_until)
                            now = _now_local()
                            if now >= until:
                                # Pause expired, clear it
                                _tracking_state['pause_until'] = None
                                _tracking_state['pause_kind'] = None
                                _save_tracking_state()
                                print(f"[Schedule] Cleared expired pause from {pause_until}")
                        except Exception:
                            # Invalid pause format, clear it
                            _tracking_state['pause_until'] = None
                            _tracking_state['pause_kind'] = None
                            _save_tracking_state()
                            print(f"[Schedule] Cleared invalid pause format: {pause_until}")
    except Exception:
        pass

def _save_tracking_state():
    try:
        with open(TRACK_STATE_PATH, 'w', encoding='utf-8') as f:
            json.dump(_tracking_state, f, indent=2)
        print(f"[Schedule] Saved tracking state to {TRACK_STATE_PATH}")
    except Exception as e:
        print(f"[ERROR] Failed to save tracking state to {TRACK_STATE_PATH}: {e}")

def _parse_range(s: str):
    try:
        a, b = [x.strip() for x in s.split('-', 1)]
        a_h, a_m = [int(x) for x in a.split(':', 1)]
        b_h, b_m = [int(x) for x in b.split(':', 1)]
        return (a_h, a_m), (b_h, b_m)
    except Exception:
        return (8,30), (17,30)

def _now_local():
    """Get current time in local timezone (WIB = UTC+7)."""
    # Return timezone-aware WIB datetime
    wib_tz = dt.timezone(dt.timedelta(hours=7))
    return dt.datetime.now(wib_tz)

def _now_wib():
    """Get current time in WIB timezone (UTC+7) - alias for _now_local()."""
    return _now_local()

def _now_utc():
    """Get current time in UTC timezone (timezone-aware)."""
    return dt.datetime.now(dt.timezone.utc)

def _in_range(now: dt.datetime, rng: str) -> bool:
    """Check if a UTC datetime falls within a local time range (HH:MM-HH:MM)."""
    (h1, m1), (h2, m2) = _parse_range(rng)
    try:
        # The 'now' parameter is the timestamp of the log event (in UTC)
        # Convert it to local time (WIB) for comparison against the schedule.
        log_time_utc = now.replace(tzinfo=dt.timezone.utc)
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        log_time_local = log_time_utc.astimezone(wib_tz)
        
        start_time = dt.time(h1, m1)
        end_time = dt.time(h2, m2)
        
        # Compare the time part of the log event with the schedule range.
        return start_time <= log_time_local.time() < end_time
    except Exception:
        return False # Default to false on any parsing/conversion error

def _maybe_update_tracking_state():
    now_utc = _now_utc()
    # Handle manual pause
    pu = _tracking_state.get('pause_until')
    pk = _tracking_state.get('pause_kind')
    if pu:
        try:
            until = dt.datetime.fromisoformat(pu) # Assume pause_until is local server time
            now_for_pause_check = _now_local()
            if now_for_pause_check < until:
                # During manual pause
                kind = (pk or '').lower()
                if kind == 'lunch':
                    # Lunch: system active but alerts suppressed
                    _tracking_state['tracking_active'] = True
                    _tracking_state['suppress_alerts'] = True
                else:
                    # Off-hours (default): system inactive, alerts not suppressed (already off)
                    _tracking_state['tracking_active'] = False
                    _tracking_state['suppress_alerts'] = False
                return
            else:
                # Pause expired
                _tracking_state['pause_until'] = None
                _tracking_state['pause_kind'] = None
        except Exception:
            _tracking_state['pause_until'] = None
            _tracking_state['pause_kind'] = None
    if bool(_tracking_state.get('auto_schedule', True)):
        work_ok = _in_range(now_utc, str(_tracking_state.get('work_hours', '08:30-17:30')))
        lunch_on = _in_range(now_utc, str(_tracking_state.get('lunch_break', '12:00-13:00')))
        _tracking_state['tracking_active'] = bool(work_ok)
        _tracking_state['suppress_alerts'] = bool(lunch_on)
    # persist periodically via caller

def schedule_tracking_manager():
    def _job():
        while True:
            try:
                _maybe_update_tracking_state()
                _save_tracking_state()
            except Exception:
                pass
            time.sleep(15)
    t = threading.Thread(target=_job, daemon=True)
    t.start()

@app.route('/api/schedule/state')
def api_schedule_state():
    _maybe_update_tracking_state()
    st = dict(_tracking_state)
    return jsonify(st)

@app.route('/api/schedule/mode', methods=['POST'])
def api_schedule_mode():
    data = request.get_json(silent=True) or {}
    auto = data.get('auto_schedule')
    if auto is not None:
        _tracking_state['auto_schedule'] = bool(auto)
    # Allow manual overrides when auto_schedule is False
    if not _tracking_state.get('auto_schedule', True):
        if 'tracking_active' in data:
            _tracking_state['tracking_active'] = bool(data.get('tracking_active'))
        if 'suppress_alerts' in data:
            _tracking_state['suppress_alerts'] = bool(data.get('suppress_alerts'))
    # Update schedule ranges if provided
    wh = data.get('work_hours')
    lb = data.get('lunch_break')
    if isinstance(wh, str) and '-' in wh:
        _tracking_state['work_hours'] = wh
    if isinstance(lb, str) and '-' in lb:
        _tracking_state['lunch_break'] = lb
    # Clear manual pause if any
    if data.get('clear_pause'):
        _tracking_state['pause_until'] = None
        _tracking_state['pause_kind'] = None
    _maybe_update_tracking_state()
    _save_tracking_state()
    return jsonify({'ok': True, 'state': _tracking_state})

@app.route('/api/schedule/pause', methods=['POST'])
def api_schedule_pause():
    data = request.get_json(silent=True) or {}
    minutes = data.get('minutes')
    until_s = data.get('until')  # ISO local time string
    kind = (data.get('kind') or '').lower()  # 'lunch' or 'offhours'
    now = _now_local()
    until = None
    if isinstance(minutes, (int, float)) and minutes > 0:
        until = now + dt.timedelta(minutes=float(minutes))
    elif isinstance(until_s, str):
        try:
            until = dt.datetime.fromisoformat(until_s)
        except Exception:
            pass
    if until is None:
        return jsonify({'error': 'invalid_pause'}), 400
    _tracking_state['pause_until'] = until.isoformat()
    _tracking_state['pause_kind'] = (kind if kind in ('lunch','offhours') else 'offhours')
    _maybe_update_tracking_state()
    _save_tracking_state()
    return jsonify({'ok': True, 'state': _tracking_state})


@app.route('/')
def index():
    return render_template('index.html')


# --- System Information API ---
def _internet_latency_ms(host: str = '8.8.8.8', port: int = 53, timeout: float = 1.0) -> Optional[float]:
    try:
        t0 = time.time()
        with socket.create_connection((host, port), timeout=timeout):
            dt_ms = (time.time() - t0) * 1000.0
            return float(dt_ms)
    except Exception:
        return None

# --- GPU Usage Function --- 
def _gpu_usage_percent() -> Optional[int]:
    """Return first GPU utilization percent if nvidia-smi is available; else None."""
    try:
        # Windows: nvidia-smi usually on PATH with driver; otherwise users can add it.
        p = subprocess.run(['nvidia-smi', '--query-gpu=utilization.gpu', '--format=csv,noheader,nounits'],
                           stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=1.5)
        if p.returncode != 0:
            return None
        out = (p.stdout or '').strip().splitlines()
        if not out:
            return None
        # take first GPU
        try:
            val = int(out[0].strip())
            return max(0, min(100, val))
        except Exception:
            return None
    except Exception:
        return None


class MEMORYSTATUSEX(ctypes.Structure):
    _fields_ = [
        ("dwLength", wintypes.DWORD),
        ("dwMemoryLoad", wintypes.DWORD),
        ("ullTotalPhys", ctypes.c_uint64),
        ("ullAvailPhys", ctypes.c_uint64),
        ("ullTotalPageFile", ctypes.c_uint64),
        ("ullAvailPageFile", ctypes.c_uint64),
        ("ullTotalVirtual", ctypes.c_uint64),
        ("ullAvailVirtual", ctypes.c_uint64),
        ("ullAvailExtendedVirtual", ctypes.c_uint64),
    ]

# --- memory usage function ---
def _memory_status_bytes() -> Optional[Dict[str, int]]:
    """Return dict with total and available physical memory in bytes (Windows-friendly)."""
    try:
        stat = MEMORYSTATUSEX()
        stat.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
        if hasattr(ctypes.windll, 'kernel32') and ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(stat)):
            total = int(stat.ullTotalPhys)
            avail = int(stat.ullAvailPhys)
            return { 'total': total, 'available': avail }
    except Exception:
        pass
    # Fallback using shutil on POSIX or when ctypes fails (approx total only)
    try:
        import psutil  # if available
        vm = psutil.virtual_memory()
        return { 'total': int(vm.total), 'available': int(vm.available) }
    except Exception:
        try:
            # As last resort, return None to avoid blocking
            return None
        except Exception:
            return None


@app.route('/api/system/info')
def api_system_info():
    lat = _internet_latency_ms()
    gpu = _gpu_usage_percent()
    mem = _memory_status_bytes() or {}
    total_b = int(mem.get('total') or 0)
    avail_b = int(mem.get('available') or 0)
    used_b = total_b - avail_b if total_b and avail_b else None
    return jsonify({
        'internet_ms': (float(lat) if lat is not None else None),
        'gpu_usage_percent': (int(gpu) if gpu is not None else None),
        'memory_total_bytes': (int(total_b) if total_b else None),
        'memory_used_bytes': (int(used_b) if used_b is not None else None),
    })

@app.route('/api/system/uptime')
def api_system_uptime():
    """Return system uptime in seconds since startup."""
    global _system_start_time
    if _system_start_time is None:
        return jsonify({'uptime_seconds': 0})
    elapsed = time.time() - _system_start_time
    return jsonify({'uptime_seconds': int(elapsed)})

@app.route('/api/system/health')
def api_system_health():
    """
    Health check endpoint for monitoring system status.
    Used by frontend to detect when server is back up after restart.
    """
    global _system_start_time
    uptime = 0
    if _system_start_time is not None:
        uptime = int(time.time() - _system_start_time)

    return jsonify({
        'status': 'ok',
        'uptime_seconds': uptime,
        'timestamp': dt.datetime.now().isoformat()
    })


# Serve GSPE logo asset
@app.route('/assets/logo')
def asset_logo():
    try:
        p = os.path.join(os.path.dirname(__file__), 'templates', 'src', 'LOGO_GSPE_transparent.png')
        return send_file(p, mimetype='image/png')
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/api/cameras')
def api_cameras():
    global _camera_map
    _camera_map = load_cameras()
    # Do not expose rtsp_url to the client; include stream_enabled for UI control
    payload = [{
        'id': c['id'],
        'name': c.get('name', f"CAM {c['id']}"),
        'stream_enabled': bool(c.get('stream_enabled', True)),
        'area': c.get('area', ''),
    } for c in _camera_map.values()]
    # sort by id for stable ordering
    payload.sort(key=lambda x: x['id'])
    return jsonify(payload)


@app.route('/api/cameras/status')
def api_cameras_status():
    """Return current AI and Stream status per camera for UI polling."""
    cams = load_cameras()
    items = []
    for cam_id in sorted(cams.keys()):
        ai_running = False
        try:
            ai_running = bool(ai_manager and hasattr(ai_manager, 'is_camera_running') and ai_manager.is_camera_running(cam_id))
        except Exception:
            ai_running = False
        items.append({
            'id': cam_id,
            'name': cams[cam_id].get('name') or f'CAM{cam_id}',
            'ai_running': ai_running,
            'stream_enabled': bool(cams[cam_id].get('stream_enabled', True)),
        })
    return jsonify({'items': items})


@app.route('/api/cameras', methods=['POST'])
def api_add_camera():
    """Create a new camera by writing camera_configs/CAM<ID>/config.json."""
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    rtsp_url = (data.get('rtsp_url') or '').strip()
    cam_id = data.get('id')
    area = (data.get('area') or '').strip()
    if not name or not rtsp_url:
        return jsonify({'error': 'name and rtsp_url are required'}), 400
    try:
        cams = load_cameras()
        # determine id
        if cam_id is None or str(cam_id).strip() == '':
            next_id = (max(cams.keys()) + 1) if cams else 1
        else:
            next_id = int(cam_id)
            if next_id in cams:
                return jsonify({'error': 'id already exists'}), 409
        # prepare folder
        folder = os.path.join(CAMERA_DIR, f'CAM{next_id}')
        cfg_path = os.path.join(folder, 'config.json')
        if os.path.exists(cfg_path):
            return jsonify({'error': 'config already exists'}), 409
        os.makedirs(folder, exist_ok=True)
        with open(cfg_path, 'w', encoding='utf-8') as f:
            json.dump({'id': next_id, 'name': name, 'rtsp_url': rtsp_url, 'enabled': False, 'stream_enabled': True, 'area': area}, f, indent=4)
        # refresh cache
        global _camera_map
        _camera_map = load_cameras()
        return jsonify({'ok': True, 'camera': {'id': next_id, 'name': name, 'area': area}}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Serve captures folder if needed
@app.route('/captures/<int:cam_id>/<path:filename>')
def serve_camera_capture(cam_id: int, filename: str):
    cam_dir = os.path.join(CAPTURES_DIR, str(cam_id))
    return send_from_directory(cam_dir, filename, as_attachment=False)

@app.route('/attendance_captures/<path:subpath>')
def serve_attendance_capture(subpath: str):
    return send_from_directory(ATT_CAPTURES_DIR, subpath, as_attachment=False)

@app.route('/api/captures/per_camera_latest', methods=['GET'])
def api_captures_per_camera_latest():
    """Return latest file per camera from captures/<cam_id>/ for preview cards."""
    _ensure_saver_started()
    cams = load_cameras()
    out = []
    for cam_id, meta in cams.items():
        cam_dir = os.path.join(CAPTURES_DIR, str(int(cam_id)))
        latest = None
        try:
            files = [x for x in os.listdir(cam_dir) if x.lower().endswith('.jpg')]
            if files:
                files.sort()
                latest = files[-1]
        except Exception:
            latest = None
        url = None
        ts_iso = None
        if latest:
            url = f'/captures/{int(cam_id)}/{latest}'
            try:
                # Parse timestamp from filename YYYYMMDD_HHMMSS.jpg as WIB (UTC+7)
                ts_part = latest.split('.')[0]
                dt_naive = dt.datetime.strptime(ts_part, '%Y%m%d_%H%M%S')
                # Add WIB timezone info since filename is in WIB
                wib_tz = dt.timezone(dt.timedelta(hours=7))
                dt_wib = dt_naive.replace(tzinfo=wib_tz)
                ts_iso = _to_iso_utc(dt_wib)
            except Exception:
                ts_iso = None
        out.append({
            'cam_id': int(cam_id),
            'area': meta.get('area') or '',
            'name': meta.get('name') or f'CAM {cam_id}',
            'url': url,
            'timestamp': ts_iso,
        })
    return jsonify(out)

def _nearest_capture_for(cam_id: int, target_ts: dt.datetime, max_delta_sec: int = 3600) -> str:
    """Return URL to nearest capture file for camera around target_ts within window; else ''."""
    try:
        cam_dir = os.path.join(CAPTURES_DIR, str(int(cam_id)))
        files = [x for x in os.listdir(cam_dir) if x.lower().endswith('.jpg')]
        if not files:
            return ''
        # Sort and find closest by filename timestamp
        # Filenames are in WIB timezone (UTC+7)
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        best_url = ''
        best_dt = None
        best_diff = None
        for fn in files:
            try:
                ts_part = fn.split('.')[0]
                fdt_naive = dt.datetime.strptime(ts_part, '%Y%m%d_%H%M%S')
                # Add WIB timezone info
                fdt = fdt_naive.replace(tzinfo=wib_tz)
                # Ensure target_ts is also timezone-aware for comparison
                if target_ts.tzinfo is None:
                    target_ts_aware = target_ts.replace(tzinfo=wib_tz)
                else:
                    target_ts_aware = target_ts.astimezone(wib_tz)
                diff = abs((fdt - target_ts_aware).total_seconds())
                if best_diff is None or diff < best_diff:
                    best_diff = diff; best_dt = fdt; best_url = f'/captures/{int(cam_id)}/{fn}'
            except Exception:
                continue
        if best_diff is not None and best_diff <= max_delta_sec:
            return best_url
        # fallback: return latest if within window is not found
        latest = sorted(files)[-1]
        return f'/captures/{int(cam_id)}/{latest}'
    except Exception:
        return ''

def _save_attendance_capture(emp_id: int, cam_id: int, ts: dt.datetime, kind: str, meta: Dict[str, Any], overwrite: bool = False):
    """Save attendance capture image and meta under attendance_captures/YYYY-MM-DD/<emp_id>/.
    kind: 'first_in' or 'last_out'
    """
    try:
        day = (ts.date() if isinstance(ts, dt.datetime) else dt.date.today()).isoformat()
        root = os.path.join(ATT_CAPTURES_DIR, day, str(int(emp_id)))
        os.makedirs(root, exist_ok=True)
        # Strict write-once guarantee for FIRST IN unless overwrite is forced
        if kind == 'first_in' and not overwrite:
            try:
                if os.path.isfile(os.path.join(root, 'first_in.jpg')):
                    try:
                        print(f"[attendance] skip first_in: already exists for emp={emp_id} day={day}")
                    except Exception:
                        pass
                    return
            except Exception:
                # If any issue checking file existence, fall through to normal save
                pass
        # Fetch snapshot
        import urllib.request
        url = f'http://127.0.0.1:5000/api/cameras/{int(cam_id)}/snapshot?annotate=1'
        with urllib.request.urlopen(url, timeout=4.0) as resp:
            img = resp.read()
        fname = 'first_in.jpg' if kind == 'first_in' else 'last_out.jpg'
        target_path = os.path.join(root, fname)
        with open(target_path, 'wb') as f:
            f.write(img)
        try:
            print(f"[attendance] saved {kind}: emp={emp_id} cam={cam_id} day={day} file={target_path}")
        except Exception:
            pass
        # Write meta.json
        meta_path = os.path.join(root, 'meta.json')
        try:
            old = {}
            if os.path.isfile(meta_path):
                with open(meta_path, 'r', encoding='utf-8') as mf:
                    old = json.load(mf) or {}
        except Exception:
            old = {}
        if kind == 'first_in':
            old['first_in'] = {
                'ts': _to_iso_utc(ts) if isinstance(ts, dt.datetime) else None,
                'cam_id': int(cam_id),
                'cam_name': meta.get('name'),
                'cam_area': meta.get('area'),
                'file': fname,
            }
        else:
            old['last_out'] = {
                'ts': _to_iso_utc(ts) if isinstance(ts, dt.datetime) else None,
                'cam_id': int(cam_id),
                'cam_name': meta.get('name'),
                'cam_area': meta.get('area'),
                'file': fname,
            }
        with open(meta_path, 'w', encoding='utf-8') as mf:
            json.dump(old, mf, indent=2)
    except Exception:
        pass

@app.route('/api/report/attendance_captures')
def api_report_attendance_captures():
    """Return First In and Last Out capture URLs for an employee and date.
    Query params: employee_id (int), date (YYYY-MM-DD).
    """
    try:
        emp_id = int(request.args.get('employee_id') or '0')
        date_s = request.args.get('date') or ''
        if emp_id <= 0 or not date_s:
            return jsonify({'error': 'employee_id and date are required'}), 400
        try:
            day = dt.datetime.strptime(date_s, '%Y-%m-%d').date()
        except Exception:
            return jsonify({'error': 'invalid date'}), 400
        with SessionLocal() as db:
            # Load attendance row for timestamps
            att = db.query(Attendance).filter(Attendance.employee_id==emp_id, Attendance.date==day).first()
            first_in_ts = att.first_in_ts if att else None
            last_out_ts = att.last_out_ts if att else None
            # Derive event camera IDs for those timestamps (nearest events around those times)
            cam_first = None; cam_last = None
            if first_in_ts is not None:
                ev = db.query(Event).filter(Event.employee_id==emp_id, Event.timestamp>=dt.datetime.combine(day, dt.time.min), Event.timestamp<=dt.datetime.combine(day, dt.time.max)).order_by(Event.timestamp.asc()).first()
                if ev: cam_first = ev.camera_id
            if last_out_ts is not None:
                ev2 = db.query(Event).filter(Event.employee_id==emp_id, Event.timestamp>=dt.datetime.combine(day, dt.time.min), Event.timestamp<=dt.datetime.combine(day, dt.time.max)).order_by(Event.timestamp.desc()).first()
                if ev2: cam_last = ev2.camera_id
        cams_meta = load_cameras()
        first_meta = cams_meta.get(int(cam_first)) if cam_first else None
        last_meta = cams_meta.get(int(cam_last)) if cam_last else None
        # Prefer dedicated attendance captures if present
        day_s = day.isoformat()
        att_dir = os.path.join(ATT_CAPTURES_DIR, day_s, str(int(emp_id)))
        first_url = None
        last_url = None
        first_cam = {'id': int(cam_first) if cam_first else None, 'name': (first_meta.get('name') if first_meta else None), 'area': (first_meta.get('area') if first_meta else None)}
        last_cam = {'id': int(cam_last) if cam_last else None, 'name': (last_meta.get('name') if last_meta else None), 'area': (last_meta.get('area') if last_meta else None)}
        try:
            if os.path.isdir(att_dir):
                meta_path = os.path.join(att_dir, 'meta.json')
                j = {}
                try:
                    with open(meta_path, 'r', encoding='utf-8') as mf:
                        j = json.load(mf) or {}
                except Exception:
                    j = {}
                if os.path.isfile(os.path.join(att_dir, 'first_in.jpg')):
                    first_url = f'/attendance_captures/{day_s}/{int(emp_id)}/first_in.jpg'
                    if j.get('first_in'):
                        if j['first_in'].get('cam_id') is not None:
                            first_cam['id'] = j['first_in']['cam_id']
                        first_cam['name'] = j['first_in'].get('cam_name') or first_cam.get('name')
                        first_cam['area'] = j['first_in'].get('cam_area') or first_cam.get('area')
                        if j['first_in'].get('ts'):
                            first_in_ts = dt.datetime.fromisoformat(j['first_in']['ts'].replace('Z',''))
                if os.path.isfile(os.path.join(att_dir, 'last_out.jpg')):
                    last_url = f'/attendance_captures/{day_s}/{int(emp_id)}/last_out.jpg'
                    if j.get('last_out'):
                        if j['last_out'].get('cam_id') is not None:
                            last_cam['id'] = j['last_out']['cam_id']
                        last_cam['name'] = j['last_out'].get('cam_name') or last_cam.get('name')
                        last_cam['area'] = j['last_out'].get('cam_area') or last_cam.get('area')
                        if j['last_out'].get('ts'):
                            last_out_ts = dt.datetime.fromisoformat(j['last_out']['ts'].replace('Z',''))
        except Exception:
            pass
        # Fallback to rolling captures if dedicated not present
        if not first_url and (cam_first and first_in_ts):
            first_url = _nearest_capture_for(cam_first, first_in_ts)
        if not last_url and (cam_last and last_out_ts):
            last_url = _nearest_capture_for(cam_last, last_out_ts)
        out = {
            'employee_id': emp_id,
            'date': date_s,
            'first_in_ts': _to_iso_utc(first_in_ts),
            'last_out_ts': _to_iso_utc(last_out_ts),
            'first_in_url': first_url,
            'last_out_url': last_url,
            'first_in_cam': first_cam,
            'last_out_cam': last_cam,
        }
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/mark_absent', methods=['POST'])
def api_admin_mark_absent():
    data = request.get_json(silent=True) or {}
    ids = data.get('employee_ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'employee_ids required'}), 400
    today = dt.date.today()
    updated = 0
    with SessionLocal() as db:
        for eid in ids:
            try:
                e = db.get(Employee, int(eid))
                if not e:
                    continue
                att = db.query(Attendance).filter(Attendance.employee_id == e.id, Attendance.date == today).first()
                # If attendance already PRESENT today, do not mark ABSENT or change is_active
                try:
                    if att is not None and str(getattr(att, 'status', '')).upper() == 'PRESENT':
                        continue
                except Exception:
                    pass
                # Mark inactive and set ABSENT status
                if e.is_active:
                    e.is_active = False
                if att is None:
                    att = Attendance(employee_id=e.id, date=today, status='ABSENT')
                    db.add(att)
                else:
                    att.status = 'ABSENT'
                updated += 1
            except Exception:
                continue
        db.commit()
    return jsonify({'ok': True, 'updated': updated})


# --- Frame Capture APIs ---
def _ensure_dir(p: str):
    try:
        os.makedirs(p, exist_ok=True)
    except Exception:
        pass


@app.route('/api/captures', methods=['POST'])
def api_capture_save():
    data = request.get_json(silent=True) or {}
    data_url = data.get('image')
    cam_id = data.get('cam_id')
    area = data.get('area') or ''
    note = data.get('note') or ''
    # Basic validation
    if not data_url or 'base64,' not in str(data_url):
        return jsonify({'error': 'image data_url required'}), 400
    try:
        b64 = data_url.split('base64,', 1)[1]
        img_bytes = base64.b64decode(b64)
    except Exception:
        return jsonify({'error': 'invalid image'}), 400
    # Prepare paths
    now = _now_local()
    ts = now.strftime('%Y%m%d_%H%M%S')
    day_dir = os.path.join(CAPTURE_DIR, now.strftime('%Y-%m-%d'))
    _ensure_dir(day_dir)
    fname = f"cap_{ts}_cam{cam_id or 'x'}.jpg"
    fpath = os.path.join(day_dir, fname)
    try:
        with open(fpath, 'wb') as f:
            f.write(img_bytes)
        # Append log
        log = {
            'timestamp': _now_local().isoformat(timespec='seconds'),
            'file': os.path.relpath(fpath, BASE_DIR).replace('\\', '/'),
            'cam_id': cam_id,
            'area': area,
            'note': note,
        }
        _ensure_dir(CAPTURE_DIR)
        with open(os.path.join(CAPTURE_DIR, 'log.jsonl'), 'a', encoding='utf-8') as lf:
            lf.write(json.dumps(log, ensure_ascii=False) + '\n')
        # add url for direct access
        log['url'] = '/' + log['file']
        try:
            # Notify all clients that a new capture is available
            socketio.emit('capture_saved', log, broadcast=True)
        except Exception:
            pass
        return jsonify({'ok': True, 'item': log})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/captures')
def api_capture_list():
    limit = 24
    try:
        limit = int(request.args.get('limit') or 24)
    except Exception:
        limit = 24
    items = []
    # Read log backwards if exists
    log_path = os.path.join(CAPTURE_DIR, 'log.jsonl')
    try:
        if os.path.isfile(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()[-(limit*2):]  # read a bit extra
            for line in reversed(lines):
                try:
                    obj = json.loads(line.strip())
                    if isinstance(obj, dict):
                        items.append(obj)
                        if len(items) >= limit:
                            break
                except Exception:
                    continue
    except Exception:
        pass
    # Fallback: list files
    if not items:
        try:
            if os.path.isdir(CAPTURE_DIR):
                for root, _, files in os.walk(CAPTURE_DIR):
                    for fn in files:
                        if fn.lower().endswith('.jpg'):
                            fp = os.path.join(root, fn)
                            items.append({'file': os.path.relpath(fp, BASE_DIR).replace('\\', '/'), 'timestamp': None, 'cam_id': None, 'area': None, 'note': ''})
            items = sorted(items, key=lambda x: x.get('timestamp') or '', reverse=True)[:limit]
        except Exception:
            items = []
    # Provide absolute URL for files
    for it in items:
        rel = it.get('file') or ''
        it['url'] = '/' + rel.replace('\\', '/')
    return jsonify(items)


@app.route('/api/captures', methods=['DELETE'])
def api_capture_delete_by_date():
    """Delete captures for a given date (YYYY-MM-DD): files under captures/<date>/ and corresponding log lines."""
    date_s = (request.args.get('date') or '').strip()
    confirm = (request.args.get('confirm') or '').strip()
    if not date_s:
        return jsonify({'error': 'date_required'}), 400
    if confirm not in ('1', 'true', 'yes', 'on'):
        return jsonify({'error': 'confirm_required'}), 400
    try:
        # validate date format
        _ = dt.datetime.strptime(date_s, '%Y-%m-%d')
    except Exception:
        return jsonify({'error': 'invalid_date'}), 400
    # Remove directory
    day_dir = os.path.join(CAPTURE_DIR, date_s)
    removed_files = 0
    if os.path.isdir(day_dir):
        try:
            # count files for info then remove dir
            for _root, _dirs, files in os.walk(day_dir):
                for fn in files:
                    if fn.lower().endswith(('.jpg', '.jpeg', '.png')):
                        removed_files += 1
            shutil.rmtree(day_dir, ignore_errors=True)
        except Exception:
            pass
    # Filter log.jsonl
    log_path = os.path.join(CAPTURE_DIR, 'log.jsonl')
    kept = []
    removed = 0
    try:
        if os.path.isfile(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        obj = json.loads(line.strip())
                        fp = (obj.get('file') or '')
                        if fp.replace('\\', '/').startswith(f'captures/{date_s}/'):
                            removed += 1
                            continue
                        kept.append(line)
                    except Exception:
                        # keep unknown lines
                        kept.append(line)
            with open(log_path, 'w', encoding='utf-8') as f:
                f.writelines(kept)
    except Exception:
        pass
    try:
        socketio.emit('captures_deleted', {'date': date_s, 'files_removed': removed_files, 'log_removed': removed}, broadcast=True)
    except Exception:
        pass
    return jsonify({'ok': True, 'date': date_s, 'files_removed': removed_files, 'log_removed': removed})


# --- Serve capture files safely ---
@app.route('/captures/<path:relpath>')
def serve_capture(relpath: str):
    try:
        # Normalize path to avoid traversal
        full = os.path.normpath(os.path.join(CAPTURE_DIR, relpath))
        if not full.startswith(os.path.normpath(CAPTURE_DIR)):
            return jsonify({'error': 'forbidden'}), 403
        if not os.path.isfile(full):
            return jsonify({'error': 'not_found'}), 404
        # Guess mimetype (most likely jpg)
        mime = 'image/jpeg'
        if full.lower().endswith('.png'):
            mime = 'image/png'
        return send_file(full, mimetype=mime)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- Camera snapshot API (works when AI inference ON; fallback to RTSP if needed) ---
def _encode_jpeg(frame) -> bytes:
    ok, buf = cv2.imencode('.jpg', frame)
    if not ok:
        raise RuntimeError('encode_failed')
    return buf.tobytes()


def _annotate_faces(frame: Optional[np.ndarray]) -> Optional[np.ndarray]:
    """Run face detection and draw bounding boxes. Return annotated frame or None if not available."""
    if frame is None:
        return None
    try:
        app_engine = _get_face_app()
        if app_engine is None:
            return frame
        faces = app_engine.get(frame)
        if not faces:
            return frame
        out = frame.copy()
        for f in faces:
            try:
                box = f.bbox.astype(int)
                x1, y1, x2, y2 = [int(v) for v in box]
                cv2.rectangle(out, (x1, y1), (x2, y2), (0, 255, 0), 2)
            except Exception:
                continue
        return out
    except Exception:
        return frame


def _apply_annotation(frame: Optional[np.ndarray], cam_id: Optional[int], annotate: bool) -> Optional[np.ndarray]:
    """Prefer ai_manager.annotate_frame for identical UI styling; fallback to _annotate_faces."""
    if frame is None or not annotate:
        return frame
    try:
        if ai_manager is not None and hasattr(ai_manager, 'annotate_frame'):
            try:
                return ai_manager.annotate_frame(frame, cam_id)
            except Exception:
                pass
        return _annotate_faces(frame)
    except Exception:
        return frame


@app.route('/api/cameras/<int:cam_id>/snapshot')
def api_camera_snapshot(cam_id: int):
    annotate = str(request.args.get('annotate') or '').strip() in ('1', 'true', 'yes', 'on')
    # 1) Try AI manager last-frame buffer, if provided by implementation
    try:
        if ai_manager is not None:
            # Common callable methods
            for attr in ['get_last_frame', 'get_snapshot', 'snapshot', 'last_frame', 'frame_for']:
                if hasattr(ai_manager, attr):
                    fn = getattr(ai_manager, attr)
                    if callable(fn):
                        try:
                            frame = fn(cam_id)
                            if isinstance(frame, np.ndarray):
                                if annotate:
                                    frame = ai_manager.annotate_frame(frame, cam_id) if hasattr(ai_manager, 'annotate_frame') else _apply_annotation(frame, cam_id, annotate)
                                data = _encode_jpeg(frame)
                                return Response(data, mimetype='image/jpeg')
                        except Exception:
                            pass
            # Inspect camera object containers
            candidates = []
            try:
                if hasattr(ai_manager, 'get_camera'):
                    c = ai_manager.get_camera(cam_id)
                    if c is not None:
                        candidates.append(c)
            except Exception:
                pass
            try:
                if hasattr(ai_manager, 'cameras'):
                    cams_attr = ai_manager.cameras
                    # mapping or list
                    if isinstance(cams_attr, dict):
                        c = cams_attr.get(cam_id)
                        if c is not None:
                            candidates.append(c)
                    elif hasattr(cams_attr, '__iter__'):
                        for c in cams_attr:
                            try:
                                if getattr(c, 'id', None) == cam_id:
                                    candidates.append(c)
                                    break
                            except Exception:
                                continue
            except Exception:
                pass
            try:
                if hasattr(ai_manager, 'streams'):
                    s = ai_manager.streams
                    if isinstance(s, dict):
                        v = s.get(cam_id)
                        if v is not None:
                            candidates.append(v)
            except Exception:
                pass
            # Probe likely attributes on candidate objects
            for obj in candidates:
                for attr in ['last_frame', 'current_frame', 'frame', 'image', 'last_image']:
                    try:
                        val = getattr(obj, attr, None)
                        if isinstance(val, np.ndarray):
                            fr = _apply_annotation(val, cam_id, annotate) or val
                            data = _encode_jpeg(fr)
                            return Response(data, mimetype='image/jpeg')
                        # Some store bytes (already encoded)
                        if isinstance(val, (bytes, bytearray)):
                            return Response(bytes(val), mimetype='image/jpeg')
                        # Some expose a callable to obtain frame
                        if callable(val):
                            fr = val()
                            if isinstance(fr, np.ndarray):
                                fr = _apply_annotation(fr, cam_id, annotate) or fr
                                data = _encode_jpeg(fr)
                                return Response(data, mimetype='image/jpeg')
                    except Exception:
                        continue
    except Exception:
        pass

    # 2) Fallback to direct RTSP/VideoCapture single-frame grab
    try:
        cams = load_cameras()
        cam = cams.get(cam_id)
        if not cam:
            return jsonify({'error': 'camera_not_found'}), 404
        rtsp = cam.get('rtsp_url') or ''
        if not rtsp:
            return jsonify({'error': 'no_rtsp'}), 404
        cap = None
        try:
            cap = cv2.VideoCapture(rtsp)
            # Try to read up to ~2s
            t0 = time.time()
            frame = None
            while time.time() - t0 < 2.0:
                ok, f = cap.read()
                if ok and f is not None:
                    frame = f
                    break
            if frame is None:
                return jsonify({'error': 'no_frame'}), 504
            frame = _apply_annotation(frame, cam_id, annotate) or frame
            data = _encode_jpeg(frame)
            return Response(data, mimetype='image/jpeg')
        finally:
            try:
                if cap is not None:
                    cap.release()
            except Exception:
                pass
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- Reports: Attendance & Alert Logs ---

# Helper function: Check if timestamp falls within time range
# Note: _in_range() function is defined earlier at line 443 with timezone-aware conversion
# This duplicate has been removed to prevent confusion


# Helper function: Batch compute violation counts to avoid N+1 query problem
def _compute_violation_counts_batch(db, attendance_rows: list) -> dict:
    """
    Compute violation counts for multiple attendance records in a single batch query.
    Returns: Dict[(employee_id, date), violation_count]

    This replaces the N+1 query pattern where we queried AlertLog for each attendance row.
    Performance: O(N) queries -> O(1) query (100x faster for large datasets)
    """
    violation_map = {}

    if not attendance_rows:
        return violation_map

    try:
        # Extract unique employee IDs and date range
        emp_ids = list(set([att.employee_id for att, emp in attendance_rows]))
        dates = [att.date for att, emp in attendance_rows if att.date]

        if not dates:
            return violation_map

        min_date = min(dates)
        max_date = max(dates)
        min_dt = dt.datetime.combine(min_date, dt.time.min)
        max_dt = dt.datetime.combine(max_date, dt.time.max)

        # Single batch query for all EXIT alerts in date range
        all_exit_logs = db.query(AlertLog).filter(
            AlertLog.employee_id.in_(emp_ids),
            AlertLog.alert_type == 'EXIT',
            AlertLog.timestamp >= min_dt,
            AlertLog.timestamp <= max_dt
        ).all()

        # Group logs by (employee_id, date) and count violations
        for log in all_exit_logs:
            log_date = log.timestamp.date()
            key = (log.employee_id, log_date)

            # Check if this EXIT qualifies as a violation
            if (bool(log.schedule_tracking_active) and
                not bool(log.schedule_is_manual_pause) and
                not _in_range(log.timestamp, log.schedule_lunch_break or '12:00-13:00')):

                violation_map[key] = violation_map.get(key, 0) + 1

    except Exception as e:
        print(f"[Violation Batch] Error computing violations: {e}")

    return violation_map


@app.route('/api/report/attendance')
def api_report_attendance():
    """Return attendance rows with optional filters. Supports JSON or Excel via ?format=xlsx."""
    args = request.args
    from_s = args.get('from') or args.get('date_from') or args.get('start')
    to_s = args.get('to') or args.get('date_to') or args.get('end')
    emp_id = args.get('employee_id')
    fmt = (args.get('format') or '').lower()

    start_d = end_d = None
    try:
        if from_s:
            y, m, d = [int(x) for x in str(from_s).split('-')]
            start_d = dt.date(y, m, d)
        if to_s:
            y, m, d = [int(x) for x in str(to_s).split('-')]
            end_d = dt.date(y, m, d)
    except Exception:
        return jsonify({'error': 'invalid_date'}), 400

    with SessionLocal() as db:
        q = db.query(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.id)
        if emp_id:
            try:
                q = q.filter(Attendance.employee_id == int(emp_id))
            except Exception:
                pass
        if start_d:
            q = q.filter(Attendance.date >= start_d)
        if end_d:
            q = q.filter(Attendance.date <= end_d)
        q = q.order_by(Attendance.date.desc(), Employee.name.asc())
        rows = q.all()

        # Excel Output
        if fmt == 'xlsx':
            # Compute violation counts in batch (avoid N+1 query problem)
            violation_map = _compute_violation_counts_batch(db, rows)

            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            headers = ['Employee Code', 'Employee Name', 'Date', 'First In', 'Last Out', 'Status', 'Violation', 'First In Camera', 'Last Out Camera']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for att, emp in rows:
                # Get violation count from batch map
                vio = 0
                if att.date:
                    vio = violation_map.get((att.employee_id, att.date), 0)

                # Read capture metadata from meta.json
                first_in_camera = ''
                last_out_camera = ''

                if att.date and att.employee_id:
                    day_s = att.date.isoformat()
                    att_dir = os.path.join(ATT_CAPTURES_DIR, day_s, str(att.employee_id))
                    meta_path = os.path.join(att_dir, 'meta.json')

                    if os.path.isfile(meta_path):
                        try:
                            with open(meta_path, 'r', encoding='utf-8') as f:
                                meta = json.load(f)

                            # Extract First In camera info
                            if meta.get('first_in'):
                                first_in = meta['first_in']
                                area = first_in.get('cam_area', '')
                                name = first_in.get('cam_name', '')
                                if area and name:
                                    first_in_camera = f"{area} - {name}"
                                elif name:
                                    first_in_camera = name
                                elif area:
                                    first_in_camera = area

                            # Extract Last Out camera info
                            if meta.get('last_out'):
                                last_out = meta['last_out']
                                area = last_out.get('cam_area', '')
                                name = last_out.get('cam_name', '')
                                if area and name:
                                    last_out_camera = f"{area} - {name}"
                                elif name:
                                    last_out_camera = name
                                elif area:
                                    last_out_camera = area
                        except Exception as e:
                            print(f"[Excel Export] Error reading meta.json for emp={att.employee_id} date={day_s}: {e}")

                # Append row data
                ws.append([
                    emp.employee_code or '',
                    emp.name or '',
                    att.date.isoformat() if att.date else '',
                    _to_wib_string(att.first_in_ts) or '',
                    _to_wib_string(att.last_out_ts) or '',
                    att.status or '',
                    vio,
                    first_in_camera,
                    last_out_camera,
                ])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return Response(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=attendance_report.xlsx'})

        # JSON output (default)
        # Use batch violation counts (same optimization as Excel)
        violation_map = _compute_violation_counts_batch(db, rows)

        data = []
        for att, emp in rows:
            # Get violation count from batch map
            vio = 0
            if att.date:
                vio = violation_map.get((att.employee_id, att.date), 0)

            data.append({
                'employee_id': att.employee_id,
                'employee_code': emp.employee_code,
                'employee_name': emp.name,
                'date': att.date.isoformat() if att.date else None,
                'first_in_ts': _to_iso_utc(att.first_in_ts),
                'last_out_ts': _to_iso_utc(att.last_out_ts),
                'status': att.status,
                'entry_type': att.entry_type or 'AUTO',
                'violation_count': vio,
            })
        return jsonify(data)


# --- Manual Attendance Entry API ---
@app.route('/api/attendance/manual', methods=['POST'])
def api_set_manual_attendance():
    """
    Allow admin to manually set attendance status for an employee on a specific date.
    Manual entries are protected from auto-override by scheduler and AI detection.

    Request JSON:
    {
        "employee_id": int,
        "date": "YYYY-MM-DD",
        "status": "PRESENT" or "ABSENT"
    }
    """
    data = request.get_json(silent=True) or {}
    emp_id = data.get('employee_id')
    date_str = data.get('date')
    status = data.get('status', 'PRESENT').upper()

    # Validation
    if not emp_id or not date_str:
        return jsonify({'error': 'employee_id and date required'}), 400

    if status not in ('PRESENT', 'ABSENT'):
        return jsonify({'error': 'status must be PRESENT or ABSENT'}), 400

    try:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'invalid date format, expected YYYY-MM-DD'}), 400

    try:
        with SessionLocal() as db:
            # Check if employee exists
            emp = db.get(Employee, int(emp_id))
            if not emp:
                return jsonify({'error': 'employee not found'}), 404

            # Upsert attendance record
            att = db.query(Attendance).filter(
                Attendance.employee_id == emp_id,
                Attendance.date == date_obj
            ).first()

            if att is None:
                # Create new manual entry
                att = Attendance(
                    employee_id=emp_id,
                    date=date_obj,
                    status=status,
                    entry_type='MANUAL',
                    first_in_ts=None,
                    last_out_ts=None
                )
                db.add(att)
            else:
                # Update existing record to manual
                att.status = status
                att.entry_type = 'MANUAL'

            db.commit()

            return jsonify({
                'ok': True,
                'attendance': {
                    'employee_id': att.employee_id,
                    'employee_name': emp.name,
                    'date': att.date.isoformat(),
                    'status': att.status,
                    'entry_type': att.entry_type,
                    'first_in_ts': _to_iso_utc(att.first_in_ts),
                    'last_out_ts': _to_iso_utc(att.last_out_ts)
                }
            })
    except Exception as e:
        print(f"[Manual Attendance] Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance/reset', methods=['POST'])
def api_reset_attendance():
    """
    Reset manual attendance entry back to AUTO mode.
    Status remains unchanged, but entry can now be auto-updated by scheduler/AI.

    Request JSON:
    {
        "employee_id": int,
        "date": "YYYY-MM-DD"
    }
    """
    data = request.get_json(silent=True) or {}
    emp_id = data.get('employee_id')
    date_str = data.get('date')

    if not emp_id or not date_str:
        return jsonify({'error': 'employee_id and date required'}), 400

    try:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'invalid date format, expected YYYY-MM-DD'}), 400

    try:
        with SessionLocal() as db:
            att = db.query(Attendance).filter(
                Attendance.employee_id == emp_id,
                Attendance.date == date_obj
            ).first()

            if not att:
                return jsonify({'error': 'attendance record not found'}), 404

            # Get employee name for response
            emp = db.get(Employee, int(emp_id))

            # Reset to AUTO mode (keep current status)
            old_type = att.entry_type
            att.entry_type = 'AUTO'
            db.commit()

            return jsonify({
                'ok': True,
                'message': 'Reset to automatic mode',
                'attendance': {
                    'employee_id': att.employee_id,
                    'employee_name': emp.name if emp else None,
                    'date': att.date.isoformat(),
                    'status': att.status,
                    'entry_type': att.entry_type,
                    'previous_type': old_type
                }
            })
    except Exception as e:
        print(f"[Reset Attendance] Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/report/alerts')
def api_report_alerts():
    """Return alert logs with optional filters. Supports JSON or Excel via ?format=xlsx."""
    args = request.args
    from_s = args.get('from') or args.get('date_from') or args.get('start')
    to_s = args.get('to') or args.get('date_to') or args.get('end')
    emp_id = args.get('employee_id')
    fmt = (args.get('format') or '').lower()
    start_dt = end_dt = None
    try:
        if from_s:
            y, m, d = [int(x) for x in str(from_s).split('-')]
            start_dt = dt.datetime(y, m, d, 0, 0, 0)
        if to_s:
            y, m, d = [int(x) for x in str(to_s).split('-')]
            end_dt = dt.datetime(y, m, d, 23, 59, 59, 999000)
    except Exception:
        return jsonify({'error': 'invalid_date'}), 400
    with SessionLocal() as db:
        q = db.query(AlertLog, Employee).join(Employee, AlertLog.employee_id == Employee.id, isouter=True)
        if emp_id:
            try:
                q = q.filter(AlertLog.employee_id == int(emp_id))
            except Exception:
                pass
        if start_dt:
            q = q.filter(AlertLog.timestamp >= start_dt)
        if end_dt:
            q = q.filter(AlertLog.timestamp <= end_dt)
        q = q.order_by(AlertLog.timestamp.desc())
        rows = q.all()

        if fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Alert Logs"
            headers = ['Timestamp', 'Employee Code', 'Employee Name', 'Alert Type', 'Message', 'Notified To']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for log, emp in rows:
                ws.append([
                    _to_wib_string(log.timestamp) or '',
                    (emp.employee_code if emp else ''),
                    (emp.name if emp else ''),
                    log.alert_type or '',
                    log.message or '',
                    log.notified_to or '',
                ])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return Response(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=alert_logs.xlsx'})

        # JSON output (default)
        data = []
        for log, emp in rows:
            data.append({
                'timestamp': _to_iso_utc(log.timestamp),
                'employee_id': log.employee_id,
                'employee_code': (emp.employee_code if emp else None),
                'employee_name': (emp.name if emp else None),
                'alert_type': log.alert_type,
                'message': log.message,
                'notified_to': log.notified_to,
            })
        return jsonify(data)

@app.route('/api/config/params')
def api_config_params():
    # Also include work hours and lunch from tracking state for convenience
    _maybe_update_tracking_state()
    return jsonify({
        'away_mute_threshold_hours': int(_app_params.get('away_mute_threshold_hours') or 15),
        'work_hours': str(_tracking_state.get('work_hours', '08:30-17:30')),
        'lunch_break': str(_tracking_state.get('lunch_break', '12:00-13:00')),
        'mark_absent_enabled': bool(_app_params.get('mark_absent_enabled', True)),
        'mark_absent_offset_minutes_before_end': int(_app_params.get('mark_absent_offset_minutes_before_end', 5)),
        'alert_min_interval_sec': int(_app_params.get('alert_min_interval_sec', 30)),
        'notification_limit': int(_app_params.get('notification_limit', 10)),
    })

# --- Alert Logs: create from client-side tracking transitions ---
@app.route('/api/alert_logs', methods=['POST'])
def api_alert_logs_create():
    try:
        payload = request.get_json(silent=True) or {}
        emp_id = payload.get('employee_id')
        alert_type = (payload.get('alert_type') or '').upper().strip()
        message = (payload.get('message') or '').strip()
        cam_id_payload = payload.get('camera_id')
        if not emp_id or alert_type not in ('ENTER','EXIT'):
            return jsonify({'error': 'invalid_payload'}), 400
        with SessionLocal() as db:
            # Optional: verify employee exists
            emp = db.get(Employee, int(emp_id))
            if not emp:
                return jsonify({'error': 'employee_not_found'}), 404
            now_utc = _now_utc()
            # Get current schedule state to snapshot it with the log
            _maybe_update_tracking_state()
            is_paused_now = bool(_tracking_state.get('pause_until'))
            is_tracking_active_now = bool(_tracking_state.get('tracking_active'))
            
            row = AlertLog(
                employee_id=int(emp_id),
                timestamp=now_utc,
                alert_type=alert_type,
                message=message or None,
                camera_id=(int(cam_id_payload) if cam_id_payload is not None else None),
                notified_to=None,
                schedule_work_hours=_tracking_state.get('work_hours'),
                schedule_lunch_break=_tracking_state.get('lunch_break'),
                schedule_is_manual_pause=is_paused_now,
                schedule_tracking_active=is_tracking_active_now,
            )
            db.add(row)

            # --- Direct Attendance Update ---
            # If this is an EXIT log, directly update today's attendance last_out_ts
            if alert_type == 'EXIT':
                today = now_utc.date()
                att = db.query(Attendance).filter(Attendance.employee_id == int(emp_id), Attendance.date == today).first()
                if att:
                    att.last_out_ts = now_utc
                else:
                    # Create attendance row if it doesn't exist for today
                    att = Attendance(employee_id=int(emp_id), date=today, last_out_ts=now_utc, status='PRESENT')
                    db.add(att)

            db.commit()
            # Notify connected clients in real-time
            try:
                socketio.emit('alert_log', {
                    'employee_id': int(emp_id),
                    'alert_type': alert_type,
                    'message': message or None,
                    'timestamp': _to_iso_utc(now_utc),
                    'employee_name': (emp.name or None),
                    'department': (emp.department or None)
                })
            except Exception:
                pass
            # Try to save dedicated attendance capture
            try:
                # Determine camera id: prefer payload, else latest Event today
                cam_id = None
                try:
                    if cam_id_payload is not None:
                        cam_id = int(cam_id_payload)
                except Exception:
                    cam_id = None
                if cam_id is None:
                    # look up nearest event for employee today
                    start_dt = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
                    end_dt = now_utc.replace(hour=23, minute=59, second=59, microsecond=999000)
                    ev_q = db.query(Event).filter(Event.employee_id==int(emp_id), Event.timestamp>=start_dt, Event.timestamp<=end_dt)
                    ev = ev_q.order_by(Event.timestamp.desc()).first()
                    if ev: cam_id = int(ev.camera_id)
                if cam_id is not None:
                    cams_meta = load_cameras(); meta = cams_meta.get(int(cam_id)) or {}
                    day_dir = os.path.join(ATT_CAPTURES_DIR, now_utc.date().isoformat(), str(int(emp_id)))
                    overwrite_first = bool(_app_params.get('attendance_first_in_overwrite_enabled', False))
                    delay_sec = 0
                    try:
                        delay_sec = int(_app_params.get('attendance_last_out_delay_sec') or 0)
                    except Exception:
                        delay_sec = 0
                    if alert_type == 'ENTER':
                        # Save first_in once, or overwrite if enabled
                        first_path = os.path.join(day_dir, 'first_in.jpg')
                        exists = os.path.isfile(first_path)
                        if (not exists) or overwrite_first:
                            _save_attendance_capture(int(emp_id), int(cam_id), now_utc, 'first_in', meta)
                    elif alert_type == 'EXIT':
                        # Optionally delay last_out capture to allow more stable frame
                        def _do_last_out_save():
                            _save_attendance_capture(int(emp_id), int(cam_id), now_utc, 'last_out', meta)
                        if delay_sec and delay_sec > 0:
                            threading.Thread(target=lambda: (time.sleep(max(0, delay_sec)), _do_last_out_save()), daemon=True).start()
                        else:
                            _do_last_out_save()
            except Exception:
                pass
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _handle_new_employee_seen(emp_id: int, cam_id: int, ts: dt.datetime):
    """
    Handles the special event when a newly registered employee is detected for the first time.
    This function is called from module_AI.
    """
    try:
        with SessionLocal() as db:
            emp = db.get(Employee, int(emp_id))
            if not emp:
                return

            # 1. Create a specific AlertLog for this event
            alert_type = 'New Employee'
            message = f"(New Employee) {emp.name} has entered the area"
            
            # Snapshot current schedule state
            _maybe_update_tracking_state()
            
            new_alert = AlertLog(
                employee_id=emp_id,
                camera_id=cam_id,
                timestamp=ts,
                alert_type=alert_type,
                message=message,
                schedule_work_hours=_tracking_state.get('work_hours'),
                schedule_lunch_break=_tracking_state.get('lunch_break'),
                schedule_is_manual_pause=bool(_tracking_state.get('pause_until')),
                schedule_tracking_active=bool(_tracking_state.get('tracking_active')),
            )
            db.add(new_alert)
            db.commit()

            # 2. Immediately trigger "First In" attendance capture
            cams_meta = load_cameras()
            meta = cams_meta.get(int(cam_id), {})
            # Force overwrite to ensure the first detection is captured, even if a placeholder existed
            _save_attendance_capture(emp_id, cam_id, ts, 'first_in', meta, overwrite=True)

            # 3. Emit a socket event to notify the web UI
            socketio.emit('alert_log', {
                'employee_id': emp_id, 'alert_type': alert_type, 'message': message,
                'timestamp': _to_iso_utc(ts), 'employee_name': emp.name, 'department': emp.department
            })
    except Exception as e:
        print(f"[ERROR] _handle_new_employee_seen failed for emp_id={emp_id}: {e}")

@app.route('/api/employees', methods=['GET'])
def list_employees():
    """Return list of employees for Manage Employee UI."""
    with SessionLocal() as db:
        rows = db.query(Employee).order_by(Employee.id.asc()).all()
        data = []
        for e in rows:
            data.append({
                'id': e.id,
                'employee_code': e.employee_code,
                'name': e.name,
                'department': e.department,
                'position': e.position,
                'phone_number': e.phone_number,
                'is_active': e.is_active,
            })
        return jsonify(data)


@app.route('/api/employees', methods=['POST'])
def add_employee():
    """Create a new employee (without face template). Face can be added via register tool."""
    payload = request.get_json(silent=True) or {}
    required = ['employee_code', 'name']
    if any(not payload.get(k) for k in required):
        return jsonify({'error': 'employee_code and name required'}), 400
    with SessionLocal() as db:
        # Check duplicate code
        exists = db.query(Employee).filter(Employee.employee_code == payload['employee_code']).first()
        if exists:
            return jsonify({'error': 'employee_code already exists'}), 409
        e = Employee(
            employee_code=payload['employee_code'],
            name=payload.get('name'),
            department=payload.get('department'),
            position=payload.get('position'),
            phone_number=payload.get('phone_number'),
            is_active=bool(payload.get('is_active', True)),
        )
        db.add(e)
        db.commit()
        return jsonify({'id': e.id}), 201


# --- Utilities ---
def _purge_attendance_captures_for_emp(emp_id: int) -> int:
    """Remove attendance_captures/<YYYY-MM-DD>/<emp_id>/ across all days. Returns count of removed day folders."""
    removed = 0
    try:
        if not os.path.isdir(ATT_CAPTURES_DIR):
            return 0
        for day in os.listdir(ATT_CAPTURES_DIR):
            day_path = os.path.join(ATT_CAPTURES_DIR, day)
            if not os.path.isdir(day_path):
                continue
            emp_dir = os.path.join(day_path, str(int(emp_id)))
            if os.path.isdir(emp_dir):
                try:
                    shutil.rmtree(emp_dir, ignore_errors=True)
                    removed += 1
                except Exception:
                    pass
    except Exception:
        pass
    return removed


# (Removed duplicate DELETE endpoint for /api/employees/<int:emp_id>)


@app.route('/api/employees/<int:eid>', methods=['PUT'])
def update_employee(eid: int):
    payload = request.get_json(silent=True) or {}
    with SessionLocal() as db:
        e = db.get(Employee, eid)
        if not e:
            return jsonify({'error': 'not found'}), 404
        # Update allowed fields
        # If updating employee_code, enforce uniqueness
        new_code = payload.get('employee_code', None)
        if new_code is not None and str(new_code) != str(e.employee_code):
            exists = db.query(Employee).filter(Employee.employee_code == new_code, Employee.id != eid).first()
            if exists:
                return jsonify({'error': 'employee_code already exists'}), 409
            e.employee_code = new_code
        for field in ['name', 'department', 'position', 'phone_number', 'is_active']:
            if field in payload:
                setattr(e, field, payload[field])
        # If is_active is toggled, update today's Attendance.status accordingly
        try:
            if 'is_active' in payload:
                today = dt.date.today()
                att = db.query(Attendance).filter(Attendance.employee_id == eid, Attendance.date == today).first()
                desired = 'PRESENT' if bool(payload.get('is_active')) else 'ABSENT'
                if att is None:
                    att = Attendance(employee_id=eid, date=today, status=desired)
                    db.add(att)
                else:
                    att.status = desired
        except Exception:
            # Do not block on attendance sync error
            pass
        db.commit()
        return jsonify({'ok': True})


@app.route('/api/employees/<int:eid>', methods=['DELETE'])
def delete_employee(eid: int):
    with SessionLocal() as db:
        e = db.get(Employee, eid)
        if not e:
            return jsonify({'error': 'not found'}), 404
        # 1) Delete dependent rows to satisfy FK constraints (explicit for safety and older DBs)
        db.query(FaceTemplate).filter(FaceTemplate.employee_id == eid).delete(synchronize_session=False)
        db.query(Attendance).filter(Attendance.employee_id == eid).delete(synchronize_session=False)
        db.query(Presence).filter(Presence.employee_id == eid).delete(synchronize_session=False)
        db.query(AlertLog).filter(AlertLog.employee_id == eid).delete(synchronize_session=False)
        # Events: remove all event rows for this employee (do not keep dangling history)
        db.query(Event).filter(Event.employee_id == eid).delete(synchronize_session=False)
        # 2) Delete employee row
        db.delete(e)
        db.commit()
        # 3) Remove face images directory from filesystem (non-fatal)
        try:
            # Remove name-based folder and legacy id-based folder
            emp_for_dir = db.get(Employee, eid)
            # emp is already deleted, so fetch name from previous object if available
            name_dir = None
            try:
                # e still holds data prior to delete
                name_dir = _employee_image_dir(e)
            except Exception:
                name_dir = None
            legacy_dir = os.path.join(FACE_IMG_DIR, str(eid))
            for path in filter(None, [name_dir, legacy_dir]):
                if os.path.isdir(path):
                    shutil.rmtree(path, ignore_errors=True)
            # Also purge attendance captures for this employee across all days
            try:
                _ = _purge_attendance_captures_for_emp(eid)
            except Exception:
                pass
        except Exception:
            pass
        return jsonify({'ok': True})


@app.route('/api/employees/<int:eid>/face_templates', methods=['POST'])
def add_face_template(eid: int):
    """Accepts a data URL image, extracts embedding with InsightFace, and stores FaceTemplate.
    Supports optional pose_label ('front'|'left'|'right'). Computes a quality_score [0..1].
    """
    payload = request.get_json(silent=True) or {}
    data_url = payload.get('image')
    pose_label = (payload.get('pose_label') or '').lower().strip() or None
    if not data_url or 'base64,' not in data_url:
        return jsonify({'error': 'image data_url required'}), 400
    app_engine = _get_face_app()
    if app_engine is None:
        return jsonify({'error': 'Face engine not available'}), 500
    # Decode data URL
    try:
        b64 = data_url.split('base64,', 1)[1]
        img_bytes = base64.b64decode(b64)
        nparr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            return jsonify({'error': 'invalid image'}), 400
    except Exception as e:
        return jsonify({'error': f'bad image: {e}'}), 400
    # Detect face and get embedding
    try:
        faces = app_engine.get(frame)
        if not faces:
            return jsonify({'error': 'no_face'}), 422
        def area(f):
            box = f.bbox.astype(int)
            return max(0, (box[2]-box[0])) * max(0, (box[3]-box[1]))
        face = max(faces, key=area)
        # Avoid boolean coercion on numpy arrays; check None explicitly
        emb = getattr(face, 'normed_embedding', None)
        if emb is None:
            emb = getattr(face, 'embedding', None)
        if emb is None:
            return jsonify({'error': 'no_embedding'}), 500
        emb = emb.astype('float32')
        emb_bytes = emb.tobytes()
        # Compute simple quality metrics
        try:
            box = face.bbox.astype(int)
            x1, y1, x2, y2 = box.tolist()
            h, w = frame.shape[:2]
            x1 = max(0, x1); y1 = max(0, y1); x2 = min(w, x2); y2 = min(h, y2)
            crop = frame[y1:y2, x1:x2]
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop is not None and crop.size > 0 else None
            # Blur score via Variance of Laplacian (normalize roughly)
            blur_var = float(cv2.Laplacian(gray, cv2.CV_64F).var()) if gray is not None else 0.0
            blur_score = max(0.0, min(1.0, blur_var / 200.0))  # 0..1
            # Brightness score
            mean_b = float(np.mean(gray)) / 255.0 if gray is not None else 0.0
            bright_score = max(0.0, min(1.0, (mean_b - 0.2) / 0.6))  # prefer 0.2..0.8
            # Face size score relative to frame
            face_area = max(1.0, float((x2 - x1) * (y2 - y1)))
            frame_area = float(max(1, w * h))
            size_score = max(0.0, min(1.0, (face_area / frame_area) / 0.1))  # 10% area -> score 1
            # Aggregate
            quality_score = float(0.5 * blur_score + 0.3 * bright_score + 0.2 * size_score)
        except Exception:
            quality_score = None
    except Exception as e:
        return jsonify({'error': f'embed_error: {e}'}), 500
    # Save
    with SessionLocal() as db:
        emp = db.get(Employee, eid)
        if not emp:
            return jsonify({'error': 'employee_not_found'}), 404
        ft = FaceTemplate(employee_id=eid, embedding=emb_bytes, pose_label=pose_label, quality_score=quality_score)
        db.add(ft)
        db.commit()

        # After we have ft.id, save cropped face image to filesystem
        try:
            box = face.bbox.astype(int)
            x1, y1, x2, y2 = box.tolist()
            # Expand a bit and clamp
            h, w = frame.shape[:2]
            pad = int(0.1 * max(x2 - x1, y2 - y1))
            x1 = max(0, x1 - pad); y1 = max(0, y1 - pad)
            x2 = min(w, x2 + pad); y2 = min(h, y2 + pad)
            crop = frame[y1:y2, x1:x2]
            if crop is not None and crop.size > 0:
                emp_dir = _employee_image_dir(emp)
                os.makedirs(emp_dir, exist_ok=True)
                out_path = os.path.join(emp_dir, f"{ft.id}.jpg")
                cv2.imwrite(out_path, crop)
        except Exception as _:
            # Non-fatal if image saving fails
            pass

        return jsonify({'ok': True, 'template_id': ft.id, 'pose_label': pose_label, 'quality_score': quality_score})


@app.route('/api/employees/<int:eid>/face_templates', methods=['GET'])
def list_face_templates(eid: int):
    """Return list of face templates for preview purposes (embedding as base64 bytes)."""
    with SessionLocal() as db:
        emp = db.get(Employee, eid)
        if not emp:
            return jsonify({'error': 'employee_not_found'}), 404
        rows = db.query(FaceTemplate).filter(FaceTemplate.employee_id == eid).order_by(FaceTemplate.id.asc()).all()
        out = []
        for r in rows:
            img_path = os.path.join(_employee_image_dir(emp), f"{r.id}.jpg")
            has_image = os.path.isfile(img_path)
            out.append({
                'id': r.id,
                'created_at': r.created_at.isoformat() if r.created_at else None,
                'pose_label': getattr(r, 'pose_label', None),
                'quality_score': getattr(r, 'quality_score', None),
                'embedding_b64': base64.b64encode(r.embedding).decode('ascii') if r.embedding else None,
                'image_url': f"/api/face_templates/{r.id}/image" if has_image else None,
            })
        return jsonify(out)


@app.route('/api/face_templates/<int:tid>/image')
def get_face_template_image(tid: int):
    """Serve the stored cropped face image for a template if available."""
    with SessionLocal() as db:
        tpl = db.get(FaceTemplate, tid)
        if not tpl:
            return jsonify({'error': 'not_found'}), 404
        emp = db.get(Employee, tpl.employee_id)
        if not emp:
            return jsonify({'error': 'employee_not_found'}), 404
        emp_dir = _employee_image_dir(emp)
    img_path = os.path.join(emp_dir, f"{tid}.jpg")
    if not os.path.isfile(img_path):
        return jsonify({'error': 'image_not_found'}), 404
    return send_file(img_path, mimetype='image/jpeg')


@app.route('/api/employees/<int:eid>/face_templates', methods=['DELETE'])
def clear_face_templates(eid: int):
    """Delete all FaceTemplate rows for employee and remove stored face crops."""
    try:
        with SessionLocal() as db:
            emp = db.get(Employee, eid)
            if not emp:
                return jsonify({'error': 'employee_not_found'}), 404
            # Delete DB rows
            db.query(FaceTemplate).filter(FaceTemplate.employee_id == eid).delete(synchronize_session=False)
            db.commit()
            # Remove images directory
            try:
                emp_dir = _employee_image_dir(emp)
                if os.path.isdir(emp_dir):
                    import shutil
                    shutil.rmtree(emp_dir, ignore_errors=True)
            except Exception:
                pass
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _check_camera_online(rtsp_url: str) -> bool:
    if cv2 is None:
        return False
    if not rtsp_url:
        return False
    try:
        src = rtsp_url.strip()
        if src.lower().startswith('webcam:'):
            idx = int(src.split(':', 1)[1])
            cap = cv2.VideoCapture(idx, getattr(cv2, 'CAP_DSHOW', 0))
        elif src.isdigit():
            cap = cv2.VideoCapture(int(src), getattr(cv2, 'CAP_DSHOW', 0))
        else:
            cap = cv2.VideoCapture(src)
        ok = bool(cap and cap.isOpened())
    except Exception:
        ok = False
    finally:
        try:
            if 'cap' in locals() and cap:
                cap.release()
        except Exception:
            pass
    return ok


def _get_camera_online_cached(cam_id: int, rtsp_url: str) -> bool:
    now = time.time()
    item = _cam_status_cache.get(cam_id)
    if item and (now - item['ts'] <= _CAM_STATUS_TTL):
        return bool(item['online'])
    online = _check_camera_online(rtsp_url)
    _cam_status_cache[cam_id] = {'ts': now, 'online': bool(online)}
    return bool(online)


class RTSPFrameSource:
    """
    Simple source that reads frames from an RTSP URL or a numeric webcam index.
    Uses OpenCV VideoCapture for both RTSP and webcam.
    """
    def __init__(self, url: str):
        self.url = url
        self._cap = None

    def open(self):
        src = int(self.url) if self.url.isdigit() else self.url
        self._cap = cv2.VideoCapture(src)
        if not self._cap or not self._cap.isOpened():
            raise RuntimeError('Failed to open video source')

    def read(self):
        if not self._cap: return None
        ok, frame = self._cap.read()
        return frame if ok else None

    def close(self):
        if self._cap: self._cap.release()

@app.route('/api/cameras/status_legacy')
def cameras_status_legacy():
    """Legacy endpoint - use /api/cameras/status instead"""
    global _camera_map
    if not _camera_map:
        _camera_map = load_cameras()
    items = []
    for cam in _camera_map.values():
        items.append({
            'id': cam['id'],
            'name': cam.get('name', f"CAM {cam['id']}")
        })
    # add online flag with cached check
    for it in items:
        rtsp_url = _camera_map[it['id']].get('rtsp_url', '')
        it['online'] = _get_camera_online_cached(it['id'], rtsp_url)
    return jsonify(items)


class StreamWorker:
    def __init__(self, sid: str, cam_id: int, rtsp_url: str):
        self.sid = sid
        self.cam_id = cam_id
        self.rtsp_url = rtsp_url
        self.stop_event = threading.Event()
        self.thread = threading.Thread(target=self.run, daemon=True)
        # Shared state (no direct capture; frames come from ai_manager)
        self.frame_lock = threading.Lock()
        # Non-blocking annotation
        self._frame_counter = 0
        self.annotate_busy = False
        self.annotate_thread: Optional[threading.Thread] = None
        self.last_annotated_frame: Optional[np.ndarray] = None

    def start(self):
        self.thread.start()

    def stop(self):
        self.stop_event.set()
        # Nothing else to stop; AI manager owns the capture

    def run(self):
        if cv2 is None:
            socketio.emit('stream_error', {'message': 'OpenCV not installed'}, to=self.sid)
            return
        if not self.rtsp_url:
            socketio.emit('stream_error', {'message': 'Invalid stream source'}, to=self.sid)
            return
        try:
            # Read AI/stream preferences once
            prefs = {
                'max_width': 960,
                'jpeg_quality': 70,
                'annotation_stride': 3,
                'target_fps': 10,
            }
            try:
                if ai_manager is not None and hasattr(ai_manager, 'get_stream_preferences'):
                    prefs.update(ai_manager.get_stream_preferences())
            except Exception:
                pass
            target_dt = 1.0 / max(1, int(prefs.get('target_fps', 10)))
            max_w = int(prefs.get('max_width', 960))
            jpeg_q = int(prefs.get('jpeg_quality', 70))
            stride = max(1, int(prefs.get('annotation_stride', 3)))
            # Determine if AI is running for this camera
            ai_running = False
            try:
                ai_running = bool(ai_manager and hasattr(ai_manager, 'is_camera_running') and ai_manager.is_camera_running(self.cam_id))
            except Exception:
                ai_running = False
            # Set up RTSP source if AI is not running (raw stream)
            src = None
            if not ai_running and RTSPFrameSource is not None:
                try:
                    src = RTSPFrameSource(self.rtsp_url)
                    src.open()
                except Exception:
                    src = None
            while not self.stop_event.is_set():
                # Source frame: from AI manager when available, otherwise from RTSP directly
                frame = None
                if ai_running:
                    try:
                        if ai_manager is not None and hasattr(ai_manager, 'get_latest_frame'):
                            frame = ai_manager.get_latest_frame(self.cam_id)
                    except Exception:
                        frame = None
                else:
                    if src is not None:
                        try:
                            frame = src.read()
                        except Exception:
                            frame = None
                if frame is None:
                    time.sleep(0.01)
                    continue
                # Downscale to reduce bandwidth/CPU
                h, w = frame.shape[:2]
                if w > max_w:
                    scale = max_w / float(w)
                    frame = cv2.resize(frame, (int(w * scale), int(h * scale)))
                # Decide annotation when AI is running; otherwise send raw frames
                frame_to_send = frame
                if ai_running:
                    self._frame_counter += 1
                    do_annotate = (self._frame_counter % stride == 0)
                    if do_annotate and not self.annotate_busy:
                        def _annotate_job(img: np.ndarray):
                            try:
                                out = img
                                if ai_manager is not None and hasattr(ai_manager, 'annotate_frame'):
                                    out = ai_manager.annotate_frame(img, self.cam_id)
                                with self.frame_lock:
                                    self.last_annotated_frame = out
                            except Exception:
                                with self.frame_lock:
                                    self.last_annotated_frame = img
                            finally:
                                self.annotate_busy = False
                        self.annotate_busy = True
                        self.annotate_thread = threading.Thread(target=_annotate_job, args=(frame.copy(),), daemon=True)
                        self.annotate_thread.start()
                    with self.frame_lock:
                        if self.last_annotated_frame is not None:
                            frame_to_send = self.last_annotated_frame
                ok, buf = cv2.imencode('.jpg', frame_to_send, [int(cv2.IMWRITE_JPEG_QUALITY), jpeg_q])
                if ok:
                    b64 = base64.b64encode(buf.tobytes()).decode('ascii')
                    socketio.emit('frame', {'image': b64, 'cam_id': self.cam_id}, to=self.sid)
                time.sleep(target_dt)
        finally:
            try:
                if 'src' in locals() and src is not None:
                    src.close()
            except Exception:
                pass
            socketio.emit('stream_stopped')


# Keep workers per client sid
_workers_by_sid: Dict[str, StreamWorker] = {}


@socketio.on('connect')
def on_connect():
    # No-op; frontend will request cameras and start a stream
    pass


@socketio.on('disconnect')
def on_disconnect():
    sid = getattr(request, 'sid', None)
    stop_worker_for_sid(sid)


# ---- Manage Camera: per-camera toggle & status ----
@socketio.on('toggle_camera')
def on_toggle_camera(data):
    try:
        cam_id = int(data.get('cam_id'))
        enable = bool(data.get('enable'))
    except Exception:
        return
    try:
        if enable:
            ai_manager.start([cam_id])
        else:
            ai_manager.stop_camera(cam_id)
    except Exception as e:
        print(f"toggle_camera error: {e}")
    # persist enabled flag to config.json
    try:
        cams = load_cameras()
        if cam_id in cams:
            folder = os.path.join(CAMERA_DIR, f'CAM{cam_id}')
            cfg_path = os.path.join(folder, 'config.json')
            if os.path.isfile(cfg_path):
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                cfg['enabled'] = bool(enable)
                with open(cfg_path, 'w', encoding='utf-8') as f:
                    json.dump(cfg, f, indent=4)
    except Exception as e:
        print(f"persist enabled failed for cam {cam_id}: {e}")
    # emit back current status
    try:
        running = ai_manager.is_camera_running(cam_id)
        socketio.emit('camera_status', {'cam_id': cam_id, 'ai_running': running})
    except Exception:
        pass


@socketio.on('toggle_ai')
def on_toggle_ai(data):
    try:
        cam_id = int(data.get('cam_id'))
        enable = bool(data.get('enable'))
    except Exception:
        return
    try:
        if enable:
            ai_manager.start([cam_id])
        else:
            ai_manager.stop_camera(cam_id)
    except Exception as e:
        print(f"toggle_ai error: {e}")
    # emit status
    try:
        running = ai_manager.is_camera_running(cam_id)
        socketio.emit('camera_status', {'cam_id': cam_id, 'ai_running': running})
    except Exception:
        pass


@socketio.on('toggle_stream')
def on_toggle_stream(data):
    try:
        cam_id = int(data.get('cam_id'))
        enable = bool(data.get('enable'))
    except Exception:
        return
    # persist to config.json
    try:
        cams = load_cameras()
        if cam_id in cams:
            folder = os.path.join(CAMERA_DIR, f'CAM{cam_id}')
            cfg_path = os.path.join(folder, 'config.json')
            if os.path.isfile(cfg_path):
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                cfg['stream_enabled'] = bool(enable)
                with open(cfg_path, 'w', encoding='utf-8') as f:
                    json.dump(cfg, f, indent=4)
    except Exception as e:
        print(f"persist stream_enabled failed for cam {cam_id}: {e}")
    # echo status
    try:
        socketio.emit('camera_status', {'cam_id': cam_id, 'stream_enabled': bool(enable)})
    except Exception:
        pass


@socketio.on('get_camera_statuses')
def on_get_camera_statuses():
    try:
        statuses = []
        cams = load_cameras()
        for cam_id in sorted(cams.keys()):
            try:
                ai_running = ai_manager.is_camera_running(cam_id)
                # auto-start if persisted enabled and not running yet
                if not ai_running and cams[cam_id].get('enabled'):
                    try:
                        ai_manager.start([cam_id])
                        ai_running = True
                    except Exception:
                        pass
            except Exception:
                ai_running = False
            statuses.append({
                'cam_id': cam_id,
                'name': cams[cam_id].get('name') or f'CAM{cam_id}',
                'ai_running': ai_running,
                'stream_enabled': bool(cams[cam_id].get('stream_enabled', True)),
            })
        socketio.emit('camera_statuses', {'items': statuses})
    except Exception as e:
        print(f"get_camera_statuses error: {e}")


def stop_worker_for_sid(sid: Optional[str]):
    if not sid:
        return
    worker = _workers_by_sid.pop(sid, None)
    if worker:
        worker.stop()


# request already imported at top


@socketio.on('start_stream')
def start_stream(payload):
    # If running in RTSP override mode, let app_rtsp.py handle this event
    try:
        if os.environ.get('APP_STREAM_MODE') == 'rtsp':
            return
    except Exception:
        pass
    sid = request.sid
    try:
        cam_id = int(payload.get('cam_id'))
    except Exception:
        emit('stream_error', {'message': 'Invalid cam_id'})
        return

    # Stop any existing stream for this client
    stop_worker_for_sid(sid)

    # Always load fresh camera config to respect latest stream toggle
    cams_now = load_cameras()
    cam = cams_now.get(cam_id)
    if not cam:
        emit('stream_error', {'message': 'Camera not found'})
        return

    # Enforce: stream must be enabled
    if not bool(cam.get('stream_enabled', True)):
        emit('stream_stopped', {'cam_id': cam_id})
        return
    # Enforce: camera AI may or may not be enabled; streaming supports both paths
    # But for annotated frames via AI path, require AI running
    try:
        # sync camera list (non-destructive)
        try:
            seed_cameras_from_configs()
        except Exception:
            pass
        is_running = bool(ai_manager and hasattr(ai_manager, 'is_camera_running') and ai_manager.is_camera_running(cam_id))
    except Exception:
        is_running = False
    # Create worker even when AI not running (raw RTSP)
    worker = StreamWorker(sid, cam_id, cam.get('rtsp_url', ''))
    _workers_by_sid[sid] = worker
    worker.start()


@socketio.on('stop_stream')
def stop_stream(payload=None):
    # If running in RTSP override mode, let app_rtsp.py handle this event
    try:
        if os.environ.get('APP_STREAM_MODE') == 'rtsp':
            return
    except Exception:
        pass
    sid = request.sid
    stop_worker_for_sid(sid)


# --- AI Tracking Endpoints ---
@app.route('/api/tracking/start', methods=['POST'])
def api_tracking_start():
    if ai_manager is None:
        return jsonify({'error': 'ai_manager_not_available'}), 500
    payload = request.get_json(silent=True) or {}
    cam_ids = payload.get('cam_ids')
    if isinstance(cam_ids, list):
        try:
            cam_ids = [int(x) for x in cam_ids]
        except Exception:
            cam_ids = None
    else:
        cam_ids = None
    try:
        # sync cameras from configs (safe if configs exist)
        seed_cameras_from_configs()
    except Exception:
        pass
    try:
        ai_manager.start(cam_ids)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tracking/stop', methods=['POST'])
def api_tracking_stop():
    if ai_manager is None:
        return jsonify({'error': 'ai_manager_not_available'}), 500
    try:
        ai_manager.stop()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/tracking/reload_embeddings', methods=['POST'])
def api_tracking_reload_embeddings():
    """Forces the AI manager to reload face embeddings from the database."""
    if ai_manager is None:
        return jsonify({'error': 'ai_manager_not_available'}), 500
    try:
        if hasattr(ai_manager, 'reload_embeddings'):
            ai_manager.reload_embeddings()
            return jsonify({'ok': True, 'message': 'Embeddings reload triggered.'})
        return jsonify({'error': 'reload_not_supported'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Pass the new employee handler to the AI manager instance
if ai_manager and hasattr(ai_manager, 'set_new_employee_callback'):
    ai_manager.set_new_employee_callback(_handle_new_employee_seen)


@app.route('/api/tracking/state')
def api_tracking_state():
    if ai_manager is None:
        # include active_total even when AI not available
        active_total = 0
        try:
            with SessionLocal() as db:
                active_total = int(db.query(Employee).filter(Employee.is_active == True).count())
        except Exception:
            active_total = 0
        return jsonify({'running': False, 'present': 0, 'alerts': 0, 'total': 0, 'active_total': active_total, 'employees': []})
    try:
        state = ai_manager.get_state()
        # Inject active_total from DB
        active_total = 0
        try:
            with SessionLocal() as db:
                active_total = int(db.query(Employee).filter(Employee.is_active == True).count())
        except Exception:
            active_total = 0
        if isinstance(state, dict):
            state = {**state, 'active_total': active_total}
        return jsonify(state)
    except Exception as e:
        return jsonify({'running': False, 'error': str(e)}), 500


# --- Admin: Reset logs by date range ---
@app.route('/api/admin/reset_logs', methods=['POST'])
def api_admin_reset_logs():
    """Delete rows in events and/or alert_logs within an optional date range.
    Request JSON: { table: 'events'|'alert_logs'|'both', from_date: 'YYYY-MM-DD' (optional), to_date: 'YYYY-MM-DD' (optional) }
    If no dates provided, delete ALL rows in selected tables.
    Returns counts deleted.
    """
    payload = request.get_json(silent=True) or {}
    table = (payload.get('table') or 'both').lower()
    from_s = payload.get('from_date')
    to_s = payload.get('to_date')
    # Parse dates to datetimes spanning whole days
    start_dt = None
    end_dt = None
    try:
        if from_s:
            y, m, d = [int(x) for x in str(from_s).split('-')]
            start_dt = dt.datetime(y, m, d, 0, 0, 0)
        if to_s:
            y, m, d = [int(x) for x in str(to_s).split('-')]
            end_dt = dt.datetime(y, m, d, 23, 59, 59, 999000)
    except Exception:
        return jsonify({'error': 'invalid_date'}), 400
    try:
        deleted_events = 0
        deleted_alerts = 0
        with SessionLocal() as db:
            if table in ('events', 'both'):
                q = db.query(Event)
                if start_dt:
                    q = q.filter(Event.timestamp >= start_dt)
                if end_dt:
                    q = q.filter(Event.timestamp <= end_dt)
                deleted_events = q.delete(synchronize_session=False)
            if table in ('alert_logs', 'both'):
                q2 = db.query(AlertLog)
                if start_dt:
                    q2 = q2.filter(AlertLog.timestamp >= start_dt)
                if end_dt:
                    q2 = q2.filter(AlertLog.timestamp <= end_dt)
                deleted_alerts = q2.delete(synchronize_session=False)
            db.commit()
        return jsonify({'ok': True, 'deleted_events': int(deleted_events), 'deleted_alert_logs': int(deleted_alerts)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- System: Restart process ---
@app.route('/api/system/restart', methods=['POST'])
def api_system_restart():
    """
    Gracefully restart the application in the same terminal.
    Uses os.execv() to replace the current process with a new one.
    """
    try:
        def _do_restart():
            try:
                # Log restart event
                print(f"[RESTART] System restart requested at {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} WIB")

                # Notify all connected Socket.IO clients that server is restarting
                try:
                    socketio.emit('server_restarting', {
                        'message': 'Server is restarting...',
                        'estimated_downtime': 10
                    }, broadcast=True)
                    socketio.sleep(0.2)  # Give time for message to be sent
                except Exception as e:
                    print(f"[RESTART] Failed to emit restart event: {e}")

                # Graceful shutdown: Stop AI manager
                try:
                    if ai_manager is not None and hasattr(ai_manager, 'stop'):
                        print("[RESTART] Stopping AI manager...")
                        ai_manager.stop()
                        print("[RESTART] AI manager stopped")
                except Exception as e:
                    print(f"[RESTART] Error stopping AI manager: {e}")

                # Graceful shutdown: Stop all stream workers
                try:
                    if _workers_by_sid:
                        print(f"[RESTART] Stopping {len(_workers_by_sid)} stream workers...")
                        for sid, worker in list(_workers_by_sid.items()):
                            try:
                                if hasattr(worker, 'stop'):
                                    worker.stop()
                            except Exception as e:
                                print(f"[RESTART] Error stopping worker {sid}: {e}")
                        print("[RESTART] Stream workers stopped")
                except Exception as e:
                    print(f"[RESTART] Error stopping stream workers: {e}")

                # Delay to allow HTTP response to flush and cleanup to complete
                print("[RESTART] Graceful shutdown completed, restarting in 1 second...")
                time.sleep(1.0)

                # Prepare restart command
                python = sys.executable
                if not python:
                    python = 'python'

                # Use os.execv() to replace current process (restart in same terminal)
                # This will NOT spawn a new terminal or background process
                args = [python] + sys.argv

                print(f"[RESTART] Executing: {' '.join(args)}")
                sys.stdout.flush()
                sys.stderr.flush()

                # Replace current process with new one (same PID, same terminal)
                # Note: os.execv() may not work on Windows - will use os.execl() as fallback
                try:
                    os.execv(python, args)
                except (OSError, AttributeError) as e:
                    print(f"[RESTART] os.execv failed ({e}), trying os.execl...")
                    os.execl(python, python, *sys.argv[1:])

            except Exception as e:
                print(f"[RESTART] Fatal error during restart: {e}")
                # Fallback to exit if execv fails
                os._exit(1)

        # Start restart in background thread to allow response to be sent first
        threading.Thread(target=_do_restart, daemon=True).start()
        return jsonify({'ok': True, 'message': 'Restarting system...'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- System: Shutdown process ---
@app.route('/api/system/shutdown', methods=['POST'])
def api_system_shutdown():
    """Gracefully stop AI/camera workers and terminate the process."""
    try:
        def _do_shutdown():
            try:
                # Stop AI manager if present
                try:
                    if ai_manager is not None and hasattr(ai_manager, 'stop'):
                        ai_manager.stop()
                except Exception:
                    pass
                # Stop all stream workers
                try:
                    for sid, worker in list(_workers_by_sid.items()):
                        try:
                            if hasattr(worker, 'stop'):
                                worker.stop()
                        except Exception:
                            pass
                except Exception:
                    pass
                time.sleep(0.3)
            finally:
                os._exit(0)
        threading.Thread(target=_do_shutdown, daemon=True).start()
        return jsonify({'ok': True, 'message': 'Shutting down...'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Initialize system start time for uptime tracking (module-level scope)
    _system_start_time = time.time()
    print(f"[STARTUP] System started at {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:

        # Load parameters once at startup
        load_params()
        # Daily maintenance: purge old events and schedule next purge
        try:
            purge_old_events()
            schedule_midnight_purge()
        except Exception:
            pass
        # Schedule absent employee detection (daily at 17:30 WIB)
        try:
            schedule_absent_detection()
        except Exception as e:
            print(f"[STARTUP] Failed to start absent detection: {e}")
        # Start background capture saver thread (5s interval)
        try:
            _start_capture_saver_thread()
        except Exception:
            pass
        # Load tracking state from file (schedule settings, work hours, etc.)
        try:
            _load_tracking_state()
            print("[STARTUP] Loaded tracking state from config/tracking_mode.json")
        except Exception as e:
            print(f"[STARTUP] Failed to load tracking state: {e}")
        # Initialize Telegram sender if enabled
        socketio.run(app, 
                    host='0.0.0.0', 
                    port=5000,
                    debug=False,
                    use_reloader=False,
                    allow_unsafe_werkzeug=True)
    except KeyboardInterrupt:
        print("\nMenghentikan server...")
        # Hentikan semua worker yang berjalan
        for sid, worker in list(_workers_by_sid.items()):
            if hasattr(worker, 'stop'):
                worker.stop()
        sys.exit(0)