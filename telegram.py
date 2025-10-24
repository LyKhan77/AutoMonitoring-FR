import os
import json
import time
import threading
from typing import Dict, Any, List, Optional
import urllib.request
import datetime as dt

# Timezone helper for WIB (UTC+7)
def _now_wib():
    """Return current datetime in WIB timezone (UTC+7)."""
    wib_tz = dt.timezone(dt.timedelta(hours=7))
    return dt.datetime.now(wib_tz)
# --- Database Connection (Isolated) ---
# This section makes telegram.py independent from app.py's models.
from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    DateTime,
    Boolean,
    Date,
    ForeignKey,
)
from sqlalchemy.orm import sessionmaker, declarative_base, relationship, joinedload

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TG_CONFIG_PATH = os.path.join(BASE_DIR, 'config', 'config_telegram.json')

# --- PostgreSQL Database Configuration (aligned with database_models.py) ---
DB_HOST = os.getenv("POSTGRES_HOST", "localhost")
DB_NAME = os.getenv("POSTGRES_DB", "FR")
DB_USER = os.getenv("POSTGRES_USER", "postgres")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD", "778899")
DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}/{DB_NAME}"

Base = declarative_base()
engine = create_engine(DATABASE_URL, echo=False, future=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine, future=True)

# --- Mirrored DB Models for telegram.py ---
class Employee(Base):
    __tablename__ = 'employees'
    id = Column(Integer, primary_key=True)
    employee_code = Column(String)
    name = Column(String)
    department = Column(String)
    is_active = Column(Boolean, default=True)

class Camera(Base):
    __tablename__ = 'cameras'
    id = Column(Integer, primary_key=True)
    name = Column(String)
    area = Column(String)

class FaceTemplate(Base):
    __tablename__ = 'face_templates'
    id = Column(Integer, primary_key=True)
    employee_id = Column(Integer, ForeignKey('employees.id'))
    pose_label = Column(String)
    # We only need these columns for finding the image path

class Attendance(Base):
    __tablename__ = 'attendances'
    id = Column(Integer, primary_key=True)
    employee_id = Column(Integer, ForeignKey('employees.id'))
    date = Column(Date)

class AlertLog(Base):
    __tablename__ = 'alert_logs'
    id = Column(Integer, primary_key=True)
    employee_id = Column(Integer, ForeignKey('employees.id'))
    timestamp = Column(DateTime, default=dt.datetime.utcnow)
    camera_id = Column(Integer, ForeignKey('cameras.id'))
    message = Column(String)
    alert_type = Column(String)
    notified_telegram = Column(Boolean, default=False, index=True)
    schedule_work_hours = Column(String)
    schedule_lunch_break = Column(String)
    schedule_is_manual_pause = Column(Boolean)
    schedule_tracking_active = Column(Boolean)
    employee = relationship("Employee", lazy="joined")
    camera = relationship("Camera")

# --- Internal State ---
_tg_cfg: Dict[str, Any] = {}
_first_start_processed = False # Flag to ensure old alerts are flushed only once
_last_known_schedule_status: Optional[str] = None # Tracks 'work_hours', 'lunch_break', 'off_hours'
_conversation_state: Dict[int, Dict[str, Any]] = {} # {chat_id: {state: 'awaiting_date', data: {}}}
_bot_active_chats = set() # Menyimpan chat_id yang sudah mengaktifkan bot
_last_update_id = 0

def _safe_name(name: Optional[str]) -> str:
    """Replicates the safe name logic from app.py for directory naming."""
    try:
        s = (name or '').strip()
        if not s:
            return 'unknown'
        cleaned = ''.join(ch for ch in s if ch.isalnum() or ch in (' ', '-', '_'))
        cleaned = '_'.join(part for part in cleaned.split())
        return cleaned[:64]
    except Exception:
        return 'unknown'

def get_latest_capture_path(cam_id: int) -> Optional[str]:
    """Get path to latest capture file for given camera ID."""
    try:
        captures_dir = os.path.join(BASE_DIR, 'captures', str(cam_id))
        if not os.path.isdir(captures_dir):
            return None

        # List all .jpg files
        files = [f for f in os.listdir(captures_dir) if f.lower().endswith('.jpg')]
        if not files:
            return None

        # Sort by filename (timestamp format: YYYYMMDD_HHMMSS.jpg)
        files.sort()
        latest_file = files[-1]

        return os.path.join(captures_dir, latest_file)
    except Exception as e:
        print(f"[get_latest_capture_path] Error for cam {cam_id}: {e}")
        return None


# --- Excel Export Helper Functions ---

def _to_wib_string(dtobj: Optional[dt.datetime]) -> Optional[str]:
    """Convert UTC datetime to WIB (UTC+7) and format as readable string."""
    if dtobj is None:
        return None
    try:
        if dtobj.tzinfo is None:
            dtobj_utc = dtobj.replace(tzinfo=dt.timezone.utc)
        else:
            dtobj_utc = dtobj.astimezone(dt.timezone.utc)
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        dtobj_wib = dtobj_utc.astimezone(wib_tz)
        return dtobj_wib.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        try:
            return dtobj.isoformat(sep=' ')
        except Exception:
            return None


def _in_range_telegram(timestamp: dt.datetime, time_range: str) -> bool:
    """Check if timestamp falls within time range (HH:MM-HH:MM)."""
    try:
        if not time_range or '-' not in time_range:
            return False
        start_s, end_s = time_range.split('-', 1)
        start_h, start_m = map(int, start_s.strip().split(':'))
        end_h, end_m = map(int, end_s.strip().split(':'))
        ts_time = timestamp.time()
        start_time = dt.time(start_h, start_m)
        end_time = dt.time(end_h, end_m)
        return start_time <= ts_time <= end_time
    except Exception:
        return False


def _compute_violation_counts_batch_telegram(db, attendance_rows: list) -> dict:
    """Compute violation counts in batch to avoid N+1 query problem."""
    from database_models import AlertLog
    violation_map = {}
    if not attendance_rows:
        return violation_map
    try:
        emp_ids = list(set([att.employee_id for att, emp in attendance_rows]))
        dates = [att.date for att, emp in attendance_rows if att.date]
        if not dates:
            return violation_map
        min_date = min(dates)
        max_date = max(dates)
        min_dt = dt.datetime.combine(min_date, dt.time.min)
        max_dt = dt.datetime.combine(max_date, dt.time.max)
        all_exit_logs = db.query(AlertLog).filter(
            AlertLog.employee_id.in_(emp_ids),
            AlertLog.alert_type == 'EXIT',
            AlertLog.timestamp >= min_dt,
            AlertLog.timestamp <= max_dt
        ).all()
        for log in all_exit_logs:
            log_date = log.timestamp.date()
            key = (log.employee_id, log_date)
            if (bool(log.schedule_tracking_active) and
                not bool(log.schedule_is_manual_pause) and
                not _in_range_telegram(log.timestamp, log.schedule_lunch_break or '12:00-13:00')):
                violation_map[key] = violation_map.get(key, 0) + 1
    except Exception as e:
        print(f"[Violation Batch] Error: {e}")
    return violation_map


def generate_excel_report(report_type: str, from_date: str, to_date: str, emp_id: Optional[int] = None) -> Optional[str]:
    """
    Generate Excel report and save to temp file.

    Args:
        report_type: 'attendance' or 'alerts'
        from_date: 'YYYY-MM-DD'
        to_date: 'YYYY-MM-DD'
        emp_id: Optional employee filter (None = all employees)

    Returns:
        Temp file path or None if failed
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from database_models import Attendance, Employee, AlertLog

        # Parse dates
        y1, m1, d1 = map(int, from_date.split('-'))
        y2, m2, d2 = map(int, to_date.split('-'))
        start_d = dt.date(y1, m1, d1)
        end_d = dt.date(y2, m2, d2)

        with SessionLocal() as db:
            if report_type == 'attendance':
                # Query attendance data
                q = db.query(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.id)
                if emp_id:
                    q = q.filter(Attendance.employee_id == emp_id)
                q = q.filter(Attendance.date >= start_d, Attendance.date <= end_d)
                q = q.order_by(Attendance.date.desc(), Employee.name.asc())
                rows = q.all()

                # Compute violation counts in batch
                violation_map = _compute_violation_counts_batch_telegram(db, rows)

                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance Report"
                headers = ['Employee Code', 'Employee Name', 'Date', 'First In', 'Last Out', 'Status', 'Violation', 'First In Camera', 'Last Out Camera']
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Populate data
                att_captures_dir = os.path.join(BASE_DIR, 'attendance_captures')
                for att, emp in rows:
                    vio = 0
                    if att.date:
                        vio = violation_map.get((att.employee_id, att.date), 0)

                    # Read capture metadata
                    first_in_camera = ''
                    last_out_camera = ''
                    if att.date and att.employee_id:
                        day_s = att.date.isoformat()
                        att_dir = os.path.join(att_captures_dir, day_s, str(att.employee_id))
                        meta_path = os.path.join(att_dir, 'meta.json')
                        if os.path.isfile(meta_path):
                            try:
                                with open(meta_path, 'r', encoding='utf-8') as f:
                                    meta = json.load(f)
                                if meta.get('first_in'):
                                    first_in = meta['first_in']
                                    area = first_in.get('cam_area', '')
                                    name = first_in.get('cam_name', '')
                                    if area and name:
                                        first_in_camera = f"{area} - {name}"
                                    elif name:
                                        first_in_camera = name
                                    elif area:
                                        first_in_camera = area
                                if meta.get('last_out'):
                                    last_out = meta['last_out']
                                    area = last_out.get('cam_area', '')
                                    name = last_out.get('cam_name', '')
                                    if area and name:
                                        last_out_camera = f"{area} - {name}"
                                    elif name:
                                        last_out_camera = name
                                    elif area:
                                        last_out_camera = area
                            except Exception:
                                pass

                    ws.append([
                        emp.employee_code or '',
                        emp.name or '',
                        att.date.isoformat() if att.date else '',
                        _to_wib_string(att.first_in_ts) or '',
                        _to_wib_string(att.last_out_ts) or '',
                        att.status or '',
                        vio,
                        first_in_camera,
                        last_out_camera,
                    ])

            elif report_type == 'alerts':
                # Query alert logs data
                start_dt = dt.datetime.combine(start_d, dt.time.min)
                end_dt = dt.datetime.combine(end_d, dt.time.max)
                q = db.query(AlertLog, Employee).join(Employee, AlertLog.employee_id == Employee.id, isouter=True)
                if emp_id:
                    q = q.filter(AlertLog.employee_id == emp_id)
                q = q.filter(AlertLog.timestamp >= start_dt, AlertLog.timestamp <= end_dt)
                q = q.order_by(AlertLog.timestamp.desc())
                rows = q.all()

                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Alert Logs"
                headers = ['Timestamp', 'Employee Code', 'Employee Name', 'Alert Type', 'Message', 'Notified To']
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Populate data
                for log, emp in rows:
                    ws.append([
                        _to_wib_string(log.timestamp) or '',
                        (emp.employee_code if emp else ''),
                        (emp.name if emp else ''),
                        log.alert_type or '',
                        log.message or '',
                        log.notified_to or '',
                    ])

            else:
                return None

        # Save to temp file
        temp_dir = os.path.join(BASE_DIR, 'temp_exports')
        os.makedirs(temp_dir, exist_ok=True)
        timestamp = _now_wib().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_{from_date}_to_{to_date}_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)
        wb.save(file_path)
        return file_path

    except Exception as e:
        print(f"[Export] Generate Excel error: {e}")
        import traceback
        traceback.print_exc()
        return None


def load_tg_config() -> Dict[str, Any]:
    """Loads the Telegram configuration from JSON file."""
    global _tg_cfg
    try:
        with open(TG_CONFIG_PATH, 'r', encoding='utf-8') as f:
            _tg_cfg = json.load(f) or {}
    except Exception as e:
        print(f"[Error] Could not load config/config_telegram.json: {e}")
        _tg_cfg = {}
    return _tg_cfg

_schedule_state_cache = {}
def _get_live_schedule_state() -> Dict[str, Any]:
    """Fetches the live schedule state from the main app's API."""
    global _schedule_state_cache
    try:
        with urllib.request.urlopen('http://127.0.0.1:5000/api/schedule/state', timeout=2.0) as resp:
            _schedule_state_cache = json.loads(resp.read().decode())
            return _schedule_state_cache
    except Exception:
        # On failure, return last known good state or a default
        return _schedule_state_cache or {'tracking_active': True, 'suppress_alerts': False}

def _is_alertable_time(ts_utc: dt.datetime) -> bool:
    """Checks if a UTC timestamp falls within the alertable work schedule (WIB)."""
    try:
        # Instead of reading a static file, get the live state from the app.
        # This respects manual pauses and auto-schedule toggles.
        schedule_state = _get_live_schedule_state()
        is_tracking_active = schedule_state.get('tracking_active', True)
        are_alerts_suppressed = schedule_state.get('suppress_alerts', False)
        # An alert should only be sent if tracking is active AND alerts are not suppressed.
        return is_tracking_active and not are_alerts_suppressed
    except Exception:
        return True # Default to true if schedule parsing fails

def send_telegram_message(chat_id: str, message: str, bot_token: str, reply_markup: Optional[Dict] = None) -> bool:
    """Sends a message to a specific Telegram chat."""
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {
        'chat_id': chat_id,
        'text': message,
        'parse_mode': 'HTML',
        'disable_web_page_preview': True,
    }
    if reply_markup:
        payload['reply_markup'] = reply_markup
    headers = {'Content-Type': 'application/json'}
    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=5.0) as resp:
            if resp.getcode() != 200:
                print(f"[Telegram] Send failed to {chat_id}: {resp.read().decode()}")
                return False
            return True
    except Exception as e:
        print(f"[Telegram] Send error to {chat_id}: {e}")
        return False

def send_telegram_photo(chat_id: str, photo_path: str, caption: str, bot_token: str) -> bool:
    """Sends a photo with a caption."""
    url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
    boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
    headers = {'Content-Type': f'multipart/form-data; boundary={boundary}'}

    parts = [
        f'--{boundary}', 'Content-Disposition: form-data; name="chat_id"', '', str(chat_id),
        f'--{boundary}', 'Content-Disposition: form-data; name="caption"', '', caption,
        f'--{boundary}', 'Content-Disposition: form-data; name="parse_mode"', '', 'HTML',
        f'--{boundary}', f'Content-Disposition: form-data; name="photo"; filename="{os.path.basename(photo_path)}"', 'Content-Type: image/jpeg', ''
    ]
    data = '\r\n'.join(parts).encode('utf-8')
    with open(photo_path, 'rb') as f:
        data += b'\r\n' + f.read() + b'\r\n'
    data += f'--{boundary}--\r\n'.encode('utf-8')

    req = urllib.request.Request(url, data=data, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=10.0) as resp:
            return 200 <= resp.getcode() < 300
    except Exception as e:
        print(f"[Telegram] Photo send error: {e}")
        return False


def send_telegram_document(chat_id: str, file_path: str, caption: str, bot_token: str) -> bool:
    """Sends a document (Excel file) with a caption."""
    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
    headers = {'Content-Type': f'multipart/form-data; boundary={boundary}'}

    parts = [
        f'--{boundary}', 'Content-Disposition: form-data; name="chat_id"', '', str(chat_id),
        f'--{boundary}', 'Content-Disposition: form-data; name="caption"', '', caption,
        f'--{boundary}', 'Content-Disposition: form-data; name="parse_mode"', '', 'HTML',
        f'--{boundary}', f'Content-Disposition: form-data; name="document"; filename="{os.path.basename(file_path)}"', 'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', ''
    ]
    data = '\r\n'.join(parts).encode('utf-8')
    with open(file_path, 'rb') as f:
        data += b'\r\n' + f.read() + b'\r\n'
    data += f'--{boundary}--\r\n'.encode('utf-8')

    req = urllib.request.Request(url, data=data, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=15.0) as resp:
            return 200 <= resp.getcode() < 300
    except Exception as e:
        print(f"[Telegram] Document send error: {e}")
        return False


def get_updates(bot_token: str, offset: int) -> List[Dict]:
    """Fetches new updates from Telegram using long polling."""
    url = f"https://api.telegram.org/bot{bot_token}/getUpdates?offset={offset}&timeout=30"
    try:
        with urllib.request.urlopen(url, timeout=35.0) as resp:
            data = json.loads(resp.read().decode())
            return data.get('result', [])
    except Exception:
        return []

def handle_callback_query(query: Dict, bot_token: str):
    """Handles presses on inline keyboard buttons."""
    chat_id = query.get('message', {}).get('chat', {}).get('id')
    data = query.get('data', '')
    state_info = _conversation_state.get(chat_id)

    if not state_info:
        return

    if state_info['state'] == 'awaiting_date' and data.startswith('date_'):
        selected_date_str = data.split('_', 1)[1]
        state_info['data']['date'] = selected_date_str
        
        # Find employees with attendance on that date
        with SessionLocal() as db:
            employees = db.query(Employee).join(Attendance).filter(
                Attendance.date == dt.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            ).order_by(Employee.name).all()

        if not employees:
            send_telegram_message(chat_id, "Tidak ada data absensi untuk tanggal tersebut.", bot_token)
            _conversation_state.pop(chat_id, None)
            return

        # Create employee selection keyboard
        keyboard = {
            "inline_keyboard": [
                [{"text": f"{emp.name}", "callback_data": f"emp_{emp.id}"}] for emp in employees
            ]
        }
        send_telegram_message(chat_id, "Silakan pilih karyawan:", bot_token, reply_markup=keyboard)
        state_info['state'] = 'awaiting_employee'

    elif state_info['state'] == 'awaiting_employee' and data.startswith('emp_'):
        emp_id = int(data.split('_', 1)[1])
        date_str = state_info['data']['date']

        # Fetch and send the attendance preview
        send_attendance_preview(chat_id, emp_id, date_str, bot_token)
        _conversation_state.pop(chat_id, None) # End conversation

    elif state_info['state'] == 'awaiting_camera_selection' and data.startswith('cam_'):
        try:
            cam_id = int(data.split('_', 1)[1])

            # Get camera info from database
            with SessionLocal() as db:
                camera = db.get(Camera, cam_id)

            if not camera:
                send_telegram_message(chat_id, "❌ Kamera tidak ditemukan.", bot_token)
                _conversation_state.pop(chat_id, None)
                return

            # Get latest capture path
            capture_path = get_latest_capture_path(cam_id)

            if not capture_path or not os.path.isfile(capture_path):
                send_telegram_message(
                    chat_id,
                    f"⚠️ Tidak ada capture tersedia untuk {camera.name or f'CAM{cam_id}'}.",
                    bot_token
                )
                _conversation_state.pop(chat_id, None)
                return

            # Extract timestamp from filename (YYYYMMDD_HHMMSS.jpg)
            filename = os.path.basename(capture_path)
            timestamp_str = filename.replace('.jpg', '').replace('.JPG', '')
            try:
                timestamp = dt.datetime.strptime(timestamp_str, '%Y%m%d_%H%M%S')
                formatted_time = timestamp.strftime('%d %b %Y, %H:%M:%S')
            except:
                formatted_time = timestamp_str

            # Format caption
            area_name = camera.area or 'Area Unknown'
            cam_name = camera.name or f'CAM{cam_id}'
            caption = f"Berikut Capture di area {area_name}-{cam_name}:\n📸 Timestamp: {formatted_time}"

            # Send photo
            success = send_telegram_photo(chat_id, capture_path, caption, bot_token)

            if not success:
                send_telegram_message(chat_id, "❌ Gagal mengirim foto capture.", bot_token)

            _conversation_state.pop(chat_id, None) # End conversation

        except Exception as e:
            print(f"[Camera Capture Callback] Error: {e}")
            send_telegram_message(chat_id, "❌ Terjadi kesalahan saat mengambil capture.", bot_token)
            _conversation_state.pop(chat_id, None)

    # Export flow callbacks
    elif state_info['state'] == 'awaiting_report_type' and data.startswith('export_'):
        report_type = data.replace('export_', '')  # 'attendance' or 'alerts'
        state_info['data']['report_type'] = report_type
        state_info['state'] = 'awaiting_employee_filter'

        # Show employee filter selection
        keyboard = {
            "inline_keyboard": [
                [{"text": "👥 All Employees", "callback_data": "filter_all"}],
                [{"text": "👤 Specific Employee", "callback_data": "filter_specific"}]
            ]
        }

        report_name = "Attendance Report" if report_type == 'attendance' else "Alert Logs"
        send_telegram_message(
            chat_id,
            f"📊 <b>{report_name}</b>\n\n👥 Filter Employee:",
            bot_token,
            reply_markup=keyboard
        )

    elif state_info['state'] == 'awaiting_employee_filter':
        if data == 'filter_all':
            # All employees - skip to date selection
            state_info['data']['employee_id'] = None
            state_info['state'] = 'awaiting_date_range'

            # Show date range presets
            keyboard = {
                "inline_keyboard": [
                    [{"text": "📅 Hari Ini", "callback_data": "range_today"}],
                    [{"text": "📅 Kemarin", "callback_data": "range_yesterday"}],
                    [{"text": "📅 7 Hari Terakhir", "callback_data": "range_7days"}],
                    [{"text": "📅 30 Hari Terakhir", "callback_data": "range_30days"}]
                ]
            }

            report_name = "Attendance Report" if state_info['data']['report_type'] == 'attendance' else "Alert Logs"
            send_telegram_message(
                chat_id,
                f"📅 <b>{report_name}</b>\n\nPilih rentang tanggal:",
                bot_token,
                reply_markup=keyboard
            )

        elif data == 'filter_specific':
            # Show employee list
            state_info['state'] = 'awaiting_employee_selection'

            try:
                with SessionLocal() as db:
                    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.name).all()

                if not employees:
                    send_telegram_message(chat_id, "⚠️ Tidak ada karyawan aktif.", bot_token)
                    _conversation_state.pop(chat_id, None)
                    return

                # Create employee selection keyboard
                keyboard = {
                    "inline_keyboard": [
                        [{"text": f"{emp.employee_code or 'EMP'} - {emp.name}", "callback_data": f"emp_{emp.id}"}]
                        for emp in employees[:20]  # Limit to 20 employees for usability
                    ]
                }

                send_telegram_message(
                    chat_id,
                    "📋 Pilih karyawan:",
                    bot_token,
                    reply_markup=keyboard
                )
            except Exception as e:
                print(f"[Export Employee List] Error: {e}")
                send_telegram_message(chat_id, "❌ Terjadi kesalahan saat memuat daftar karyawan.", bot_token)
                _conversation_state.pop(chat_id, None)

    elif state_info['state'] == 'awaiting_employee_selection' and data.startswith('emp_'):
        try:
            emp_id = int(data.split('_', 1)[1])
            state_info['data']['employee_id'] = emp_id
            state_info['state'] = 'awaiting_date_range'

            # Get employee name for confirmation
            with SessionLocal() as db:
                employee = db.get(Employee, emp_id)
                emp_name = employee.name if employee else f"ID {emp_id}"

            # Show date range presets
            keyboard = {
                "inline_keyboard": [
                    [{"text": "📅 Hari Ini", "callback_data": "range_today"}],
                    [{"text": "📅 Kemarin", "callback_data": "range_yesterday"}],
                    [{"text": "📅 7 Hari Terakhir", "callback_data": "range_7days"}],
                    [{"text": "📅 30 Hari Terakhir", "callback_data": "range_30days"}]
                ]
            }

            report_name = "Attendance Report" if state_info['data']['report_type'] == 'attendance' else "Alert Logs"
            send_telegram_message(
                chat_id,
                f"📅 <b>{report_name}</b>\n👤 Employee: {emp_name}\n\nPilih rentang tanggal:",
                bot_token,
                reply_markup=keyboard
            )
        except Exception as e:
            print(f"[Export Employee Select] Error: {e}")
            send_telegram_message(chat_id, "❌ Terjadi kesalahan.", bot_token)
            _conversation_state.pop(chat_id, None)

    elif state_info['state'] == 'awaiting_date_range' and data.startswith('range_'):
        try:
            range_type = data.replace('range_', '')
            today = dt.date.today()

            if range_type == 'today':
                from_date = to_date = today.isoformat()
            elif range_type == 'yesterday':
                yesterday = today - dt.timedelta(days=1)
                from_date = to_date = yesterday.isoformat()
            elif range_type == '7days':
                from_date = (today - dt.timedelta(days=7)).isoformat()
                to_date = today.isoformat()
            elif range_type == '30days':
                from_date = (today - dt.timedelta(days=30)).isoformat()
                to_date = today.isoformat()
            else:
                send_telegram_message(chat_id, "❌ Rentang tanggal tidak valid.", bot_token)
                _conversation_state.pop(chat_id, None)
                return

            state_info['data']['from_date'] = from_date
            state_info['data']['to_date'] = to_date

            # Generate and send report
            _generate_and_send_export(chat_id, state_info['data'], bot_token)
            _conversation_state.pop(chat_id, None)

        except Exception as e:
            print(f"[Export Date Range] Error: {e}")
            send_telegram_message(chat_id, "❌ Terjadi kesalahan.", bot_token)
            _conversation_state.pop(chat_id, None)

def _generate_and_send_export(chat_id: str, data: dict, bot_token: str):
    """Generate Excel and send to user."""
    try:
        report_type = data['report_type']
        from_date = data['from_date']
        to_date = data['to_date']
        emp_id = data.get('employee_id')

        # Get employee name if specific
        emp_name = "All Employees"
        if emp_id:
            try:
                with SessionLocal() as db:
                    employee = db.get(Employee, emp_id)
                    if employee:
                        emp_name = f"{employee.employee_code or 'EMP'} - {employee.name}"
            except Exception:
                pass

        # Show progress
        send_telegram_message(chat_id, "⏳ Generating report...", bot_token)

        # Generate Excel
        file_path = generate_excel_report(report_type, from_date, to_date, emp_id)

        if not file_path or not os.path.isfile(file_path):
            send_telegram_message(chat_id, "❌ Gagal membuat report.", bot_token)
            return

        # Count rows (for caption)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            row_count = ws.max_row - 1  # -1 for header
            wb.close()
        except Exception:
            row_count = 0

        # Format caption
        report_name = "Attendance Report" if report_type == 'attendance' else "Alert Logs"
        caption = f"✅ <b>{report_name}</b>\n👤 {emp_name}\n📅 {from_date} to {to_date}\n📊 Total: {row_count} rows"

        # Send document
        success = send_telegram_document(chat_id, file_path, caption, bot_token)

        if not success:
            send_telegram_message(chat_id, "❌ Gagal mengirim file.", bot_token)

        # Cleanup temp file
        try:
            os.remove(file_path)
        except Exception:
            pass

    except Exception as e:
        print(f"[Export Send] Error: {e}")
        import traceback
        traceback.print_exc()
        send_telegram_message(chat_id, f"❌ Terjadi kesalahan: {str(e)}", bot_token)

def send_attendance_preview(chat_id: int, emp_id: int, date_str: str, bot_token: str):
    """Fetches and sends the final attendance preview with images."""
    capture_dir = os.path.join(BASE_DIR, 'attendance_captures', date_str, str(emp_id))
    meta_path = os.path.join(capture_dir, 'meta.json')

    if not os.path.isdir(capture_dir) or not os.path.isfile(meta_path):
        send_telegram_message(chat_id, "Data capture tidak ditemukan untuk karyawan dan tanggal ini.", bot_token)
        return

    with open(meta_path, 'r') as f:
        meta = json.load(f)

    with SessionLocal() as db:
        emp = db.get(Employee, emp_id)
        emp_name = emp.name if emp else "Unknown"
        emp_code = emp.employee_code if emp else "N/A"

    send_telegram_message(chat_id, f"Berikut attendance report dari {emp_code} - {emp_name}:", bot_token)

    # First In
    first_in_data = meta.get('first_in')
    if first_in_data:
        # Correctly parse UTC timestamp and convert to local time (WIB)
        ts_utc = dt.datetime.fromisoformat(first_in_data['ts'].replace('Z', '+00:00'))
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        ts_wib = ts_utc.astimezone(wib_tz)
        ts = ts_wib.strftime('%H:%M:%S')
        cam_str = f"{first_in_data.get('cam_name', '')} - {first_in_data.get('cam_area', '')}"
        caption = f"🟢 <b>First In</b>\nTimestamp: {ts}\nCamera: {cam_str}"
        img_path = os.path.join(capture_dir, first_in_data['file'])
        if os.path.isfile(img_path):
            send_telegram_photo(chat_id, img_path, caption, bot_token)
        else:
            send_telegram_message(chat_id, caption + "\n(Capture image not found)", bot_token)

    # Last Out
    last_out_data = meta.get('last_out')
    if last_out_data:
        # Correctly parse UTC timestamp and convert to local time (WIB)
        ts_utc = dt.datetime.fromisoformat(last_out_data['ts'].replace('Z', '+00:00'))
        wib_tz = dt.timezone(dt.timedelta(hours=7))
        ts_wib = ts_utc.astimezone(wib_tz)
        ts = ts_wib.strftime('%H:%M:%S')
        cam_str = f"{last_out_data.get('cam_name', '')} - {last_out_data.get('cam_area', '')}"
        caption = f"🔴 <b>Last Out</b>\nTimestamp: {ts}\nCamera: {cam_str}"
        img_path = os.path.join(capture_dir, last_out_data['file'])
        if os.path.isfile(img_path):
            send_telegram_photo(chat_id, img_path, caption, bot_token)
        else:
            send_telegram_message(chat_id, caption + "\n(Capture image not found)", bot_token)

def process_updates(bot_token: str):
    """Handles incoming commands like /start."""
    global _last_update_id, _bot_active_chats
    
    updates = get_updates(bot_token, _last_update_id + 1)
    for update in updates:
        _last_update_id = update.get('update_id', _last_update_id)

        if 'callback_query' in update:
            handle_callback_query(update['callback_query'], bot_token)
            continue

        message = update.get('message') or update.get('edited_message')
        if not message:
            continue

        chat_id = message.get('chat', {}).get('id')
        text = message.get('text', '')

        command = text.strip().lower()

        if command.startswith('/start'):
            if chat_id not in _bot_active_chats:
                # Reset any lingering conversation state
                _conversation_state.pop(chat_id, None)
                global _first_start_processed
                if not _first_start_processed:
                    # On the very first /start, mark all old alerts as "sent"
                    print("First /start received. Flushing old alert logs...")
                    try:
                        with SessionLocal() as db:
                            db.query(AlertLog).filter(AlertLog.notified_telegram == False).update({'notified_telegram': True})
                            db.commit()
                        _first_start_processed = True
                        print("Old alerts flushed. Bot is now live.")
                    except Exception as e:
                        print(f"[Error] Failed to flush old alerts: {e}")

                print(f"Bot activated in chat ID: {chat_id}")
                _bot_active_chats.add(chat_id)
                response_message = "Haloo..👋🏼\nGSPE Monitoring Bot sudah aktif dan siap memantau laporan."
                send_telegram_message(chat_id, response_message, bot_token)
        elif command.startswith('/stop'):
            if chat_id in _bot_active_chats:
                _conversation_state.pop(chat_id, None)
                print(f"Bot paused in chat ID: {chat_id}")
                _bot_active_chats.discard(chat_id)
                response_message = "Bot telah di-Jeda ⏳"
                send_telegram_message(chat_id, response_message, bot_token)

        elif command.startswith('/help'):
            help_text = """<b>📖 Bot Command Guide</b>

<b>/start</b> - Activate bot notifications
<b>/stop</b> - Pause bot notifications
<b>/help</b> - Show this help message
<b>/status</b> - Show system status
<b>/attendance</b> - View daily attendance report
<b>/capture</b> - Get latest camera capture
<b>/export_data</b> - Export Excel reports

<b>📊 Export Data Flow:</b>
1. Choose report type (Attendance/Alerts)
2. Filter by employee (All/Specific)
3. Select date range (Today/Yesterday/7d/30d)
4. Receive Excel file

Need help? Contact your supervisor."""
            send_telegram_message(chat_id, help_text, bot_token)

        elif command.startswith('/status'):
            try:
                with SessionLocal() as db:
                    # Count active employees
                    total_emp = db.query(Employee).filter(Employee.is_active == True).count()

                    # Count attendance today
                    today = dt.date.today()
                    today_att = db.query(Attendance).filter(Attendance.date == today).count()

                    # Count alerts today
                    start_dt = dt.datetime.combine(today, dt.time.min)
                    end_dt = dt.datetime.combine(today, dt.time.max)
                    today_alerts = db.query(AlertLog).filter(
                        AlertLog.timestamp >= start_dt,
                        AlertLog.timestamp <= end_dt
                    ).count()

                    # Get camera count
                    total_cameras = db.query(Camera).count()

                wib_now = dt.datetime.now(dt.timezone(dt.timedelta(hours=7)))
                status_text = f"""<b>📊 System Status</b>

👥 <b>Employees:</b> {total_emp} active
📸 <b>Cameras:</b> {total_cameras} total
📅 <b>Attendance Today:</b> {today_att} records
🚨 <b>Alerts Today:</b> {today_alerts} logs

<i>Last updated: {wib_now.strftime('%H:%M:%S WIB')}</i>"""
                send_telegram_message(chat_id, status_text, bot_token)
            except Exception as e:
                print(f"[Status Command] Error: {e}")
                send_telegram_message(chat_id, "❌ Gagal mengambil status sistem.", bot_token)
        elif command.startswith('/attendance'):
            # Start the attendance report flow
            _conversation_state[chat_id] = {'state': 'awaiting_date', 'data': {}}
            
            today = dt.date.today()
            yesterday = today - dt.timedelta(days=1)
            day_before = today - dt.timedelta(days=2)

            keyboard = {
                "inline_keyboard": [
                    [{"text": f"Hari Ini ({today.strftime('%d %b')})", "callback_data": f"date_{today.isoformat()}"}],
                    [{"text": f"Kemarin ({yesterday.strftime('%d %b')})", "callback_data": f"date_{yesterday.isoformat()}"}],
                    [{"text": f"2 Hari Lalu ({day_before.strftime('%d %b')})", "callback_data": f"date_{day_before.isoformat()}"}]
                ]
            }
            send_telegram_message(chat_id, "Silakan pilih tanggal untuk laporan absensi:", bot_token, reply_markup=keyboard)
        elif command.startswith('/capture'):
            # Start the camera capture flow
            try:
                with SessionLocal() as db:
                    cameras = db.query(Camera).all()

                if not cameras:
                    send_telegram_message(chat_id, "⚠️ Tidak ada kamera yang tersedia.", bot_token)
                    return

                # Set conversation state
                _conversation_state[chat_id] = {'state': 'awaiting_camera_selection', 'data': {}}

                # Create inline keyboard with camera list
                keyboard = {
                    "inline_keyboard": [
                        [{"text": f"{cam.area or 'Area'} - {cam.name or f'CAM{cam.id}'}", "callback_data": f"cam_{cam.id}"}]
                        for cam in cameras
                    ]
                }

                send_telegram_message(chat_id, "📷 Silakan pilih kamera untuk melihat capture terakhir:", bot_token, reply_markup=keyboard)
            except Exception as e:
                print(f"[Capture Command] Error: {e}")
                send_telegram_message(chat_id, "❌ Terjadi kesalahan saat memuat daftar kamera.", bot_token)

        elif command.startswith('/export_data'):
            # Start the export data flow
            try:
                # Initialize conversation state
                _conversation_state[chat_id] = {'state': 'awaiting_report_type', 'data': {}}

                # Show report type selection
                keyboard = {
                    "inline_keyboard": [
                        [{"text": "📊 Attendance Report", "callback_data": "export_attendance"}],
                        [{"text": "🚨 Alert Logs", "callback_data": "export_alerts"}]
                    ]
                }

                send_telegram_message(
                    chat_id,
                    "📊 <b>Export Data Report</b>\n\nSilakan pilih jenis laporan yang ingin diunduh:",
                    bot_token,
                    reply_markup=keyboard
                )
            except Exception as e:
                print(f"[Export Command] Error: {e}")
                send_telegram_message(chat_id, "❌ Terjadi kesalahan.", bot_token)

def poll_and_send_alerts(bot_token: str):
    """Periodically checks the database for new alerts and sends them."""
    if not _bot_active_chats:
        return # Do nothing if no chat has activated the bot

    with SessionLocal() as db:
        # Ambil 10 log terbaru yang belum dikirim, dan muat data employee terkait
        alerts_to_send = db.query(AlertLog).filter(
            AlertLog.notified_telegram == False
        ).options(
            joinedload(AlertLog.employee),
            joinedload(AlertLog.camera)
        ).order_by(AlertLog.timestamp.asc()).limit(10).all()

        if not alerts_to_send:
            return

        print(f"Found {len(alerts_to_send)} new alert(s) to send.")

        for log in alerts_to_send:
            emp = log.employee
            emp_name = emp.name if emp else f"Emp ID {log.employee_id}"
            emp_code = emp.employee_code if emp else "N/A"
            dept = emp.department if emp else "-"
            
            # Format waktu ke zona waktu lokal (asumsi server di WIB)
            try:
                # Timestamp dari DB adalah UTC, konversi ke WIB (UTC+7)
                log_time_utc = log.timestamp.replace(tzinfo=dt.timezone.utc)
                wib_tz = dt.timezone(dt.timedelta(hours=7))
                log_time_wib = log_time_utc.astimezone(wib_tz)
                date_part = log_time_wib.strftime('%d-%m-%Y')
                time_part = log_time_wib.strftime('%H:%M:%S')
            except Exception:
                date_part = log.timestamp.strftime('%d-%m-%Y')
                time_part = log.timestamp.strftime('%H:%M:%S')

            # Build camera string with name and area
            cam_name = log.camera.name if log.camera and log.camera.name else None
            cam_area = log.camera.area if log.camera and log.camera.area else None
            cam_str = "Unknown"
            if cam_name and cam_area:
                cam_str = f"{cam_name} - {cam_area}"
            elif cam_name or cam_area:
                cam_str = cam_name or cam_area

            # Format pesan sesuai permintaan
            if log.alert_type == 'ENTER':
                # Pesan asli: "John Doe back to area after 5 min"
                # Kita ambil bagian "after 5 min"
                duration_part = ""
                if log.message:
                    try: duration_part = log.message.split('back to area', 1)[-1].strip()
                    except Exception: pass
                tg_msg = f"=== ALERT ===\n🟢 <b>ENTER:</b> {emp_code} - {emp_name} back to area {duration_part}\n\n<b>Department:</b> {dept}\n<b>Camera:</b> {cam_str}\n<b>Date:</b> {date_part}\n<b>Time:</b> {time_part} WIB"
            elif log.alert_type == 'New Employee':
                # Build the text message first
                caption = f"=== ALERT ===\nNew employee has entered the area. Welcome {emp_code} - {emp_name}\n\n<b>Department:</b> {dept}\n<b>Camera:</b> {cam_str}\n<b>Date:</b> {date_part}\n<b>Time:</b> {time_part} WIB"
                
                # Try to find and send a photo
                photo_sent = False
                try:
                    # Find the 'front' pose template for this employee
                    front_template = db.query(FaceTemplate).filter(
                        FaceTemplate.employee_id == log.employee_id,
                        FaceTemplate.pose_label == 'front'
                    ).order_by(FaceTemplate.id.desc()).first()

                    if front_template and emp:
                        img_dir = os.path.join(BASE_DIR, 'face_images', _safe_name(emp.name))
                        photo_path = os.path.join(img_dir, f"{front_template.id}.jpg")
                        if os.path.isfile(photo_path):
                            for chat_id in list(_bot_active_chats):
                                send_telegram_photo(chat_id, photo_path, caption, bot_token)
                            photo_sent = True
                except Exception as e:
                    print(f"[Telegram] Error finding photo for new employee: {e}")
                tg_msg = caption if not photo_sent else None # Set msg to None if photo was sent
            else: # EXIT
                duration_part = ""
                if log.message:
                    try: duration_part = log.message.split('out of area', 1)[-1].strip()
                    except Exception: pass
                tg_msg = f"=== ALERT ===\n🔴 <b>EXIT:</b> {emp_code} - {emp_name} out of area {duration_part}\n\n<b>Department:</b> {dept}\n<b>Camera:</b> {cam_str}\n<b>Date:</b> {date_part}\n<b>Time:</b> {time_part} WIB"

            # Kirim ke semua chat yang aktif
            # --- Alert Smartly Logic ---
            if tg_msg: # Only send text message if tg_msg is not None
                if _is_alertable_time(log.timestamp):
                    for chat_id in list(_bot_active_chats):
                        send_telegram_message(chat_id, tg_msg, bot_token)
                        time.sleep(0.1) # Jeda kecil antar pengiriman
            
            # Tandai sebagai sudah terkirim
            log.notified_telegram = True
        
        db.commit()

def listener_thread(bot_token: str):
    """Thread untuk mendengarkan perintah /start."""
    print("[Telegram Listener] Started. Waiting for /start command...")
    while True:
        try:
            process_updates(bot_token)
        except Exception as e:
            print(f"[Telegram Listener] Error: {e}")
            time.sleep(5) # Tunggu sebelum mencoba lagi jika ada error

def poller_thread(bot_token: str):
    """Thread untuk memeriksa dan mengirim notifikasi dari DB."""
    print("[Telegram Poller] Started. Will send alerts after activation.")
    while True:
        try:
            poll_and_send_alerts(bot_token)
        except Exception as e:
            print(f"[Telegram Poller] Error: {e}")
        time.sleep(5) # Periksa database setiap 5 detik

def main():
    """Fungsi utama untuk menjalankan bot Telegram."""
    print("Starting Telegram Bot...")
    config = load_tg_config()

    if not config.get('enabled'):
        print("Telegram bot is disabled in config/config_telegram.json. Exiting.")
        return

    bot_token = config.get('bot_token')
    if not bot_token or bot_token == 'YOUR_TELEGRAM_BOT_TOKEN':
        print("[Error] bot_token is not set in config/config_telegram.json. Exiting.")
        return

    # Jalankan listener dan poller di thread terpisah
    listener = threading.Thread(target=listener_thread, args=(bot_token,), daemon=True)
    poller = threading.Thread(target=poller_thread, args=(bot_token,), daemon=True)

    listener.start()
    poller.start()

    # Keep the main thread alive
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopping Telegram Bot.")

def _get_status_string(state: Dict[str, Any]) -> str:
    """Determines a simple status string from the schedule state object."""
    if not state.get('tracking_active', False):
        return 'off_hours'
    if state.get('suppress_alerts', False):
        return 'lunch_break'
    return 'work_hours'

def schedule_monitor_thread(bot_token: str):
    """Periodically checks for schedule status changes and notifies active chats."""
    global _last_known_schedule_status
    print("[Schedule Monitor] Started.")
    
    # On first run, just set the initial state without notifying
    try:
        initial_state = _get_live_schedule_state()
        _last_known_schedule_status = _get_status_string(initial_state)
        print(f"[Schedule Monitor] Initial status is '{_last_known_schedule_status}'")
    except Exception:
        pass

    while True:
        # Check every 5 seconds for better responsiveness to manual actions
        time.sleep(5)
        try:
            current_state = _get_live_schedule_state()
            if not current_state: # Skip if API is down
                continue

            current_status = _get_status_string(current_state)

            if current_status != _last_known_schedule_status:
                print(f"[Schedule Monitor] Status changed from '{_last_known_schedule_status}' to '{current_status}'. Notifying...")
                message = ""
                if current_status == 'lunch_break':
                    message = "=== Status Update ===\nSchedule system sedang <b>Lunch Break</b> 🍽 \n\nSelamat makan siang.."
                elif current_status == 'off_hours':
                    message = "=== Status Update ===\nSchedule system sedang <b>Off Hours</b> 🌙 \n\nSampai jumpa.."
                elif current_status == 'work_hours':
                    message = "=== Status Update ===\nSchedule system sedang <b>Work Hours</b> 💼 \n\nSaatnya kembali bekerja.."
                
                if message:
                    for chat_id in list(_bot_active_chats):
                        send_telegram_message(chat_id, message, bot_token)
                
                _last_known_schedule_status = current_status
        except Exception as e:
            print(f"[Schedule Monitor] Error: {e}")

if __name__ == '__main__':
    # Load config once at the start
    config = load_tg_config()
    bot_token = config.get('bot_token')
    if bot_token and bot_token != 'YOUR_TELEGRAM_BOT_TOKEN':
        # Start the schedule monitor thread alongside the others
        schedule_monitor = threading.Thread(target=schedule_monitor_thread, args=(bot_token,), daemon=True)
        schedule_monitor.start()
    
    # Call the main function which handles other threads and the main loop
    main()